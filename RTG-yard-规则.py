import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.font_manager as fm
import time
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional, Set
import json
import warnings
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("警告: 无法导入openpyxl，将生成示例数据")
    load_workbook = None

# 导入yard.py的功能
from yard import generate_yard_model, yard_coords, analyze_track_distribution, TRACK_WIDTH, TRACK_SPACING

# 设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
except:
    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
warnings.filterwarnings('ignore', category=UserWarning)


# ================================
# 1. 数据结构定义
# ================================
@dataclass
class BlockInfo:
    """箱区信息数据类"""
    block_id: str
    area: str
    number: int
    position_x: int
    position_y: int
    capacity: int
    current_usage: int
    has_power: bool
    track_connections: List[str]
    bay_count: int
    row_count: int
    tier_count: int


@dataclass
class TrackInfo:
    """轨道信息数据类"""
    track_id: str
    track_type: str
    start_point: Tuple[int, int]
    end_point: Tuple[int, int]
    connected_blocks: List[str]
    track_width: float
    max_rtg_count: int
    current_rtg_list: List[int]


@dataclass
class TaskInfo:
    """任务信息数据类"""
    task_id: str
    block_id: str
    bay: int
    row: int
    tier: int
    prep_time: str
    task_type: str
    truck_id: Optional[str] = None
    is_cold_container: bool = False
    manual_mark: bool = False
    estimated_duration: float = 0.0


@dataclass
class RTGInfo:
    """轮胎吊信息数据类"""
    rtg_id: int
    rtg_type: str
    current_track_id: str
    current_position: Tuple[int, int]
    login_status: str
    fault_status: str
    move_speed_horizontal: float
    move_speed_vertical: float
    operation_speed: float
    available_areas: List[str]
    max_lift_height: int
    current_load: Optional[str] = None


@dataclass
class TruckInfo:
    """拖车信息数据类"""
    truck_id: str
    current_block_id: str
    current_bay: int
    speed: float
    containers: List[str]
    arrival_time: float
    route_plan: List[str] = None


# ================================
# 2. 系统配置类
# ================================
class RealYardConfig:
    """系统配置"""
    POPULATION_SIZE = 150
    MAX_GENERATIONS = 300
    CROSSOVER_RATE = 0.9
    MUTATION_RATE = 0.25
    ELITE_RATE = 0.1
    NUM_CRANES = 12
    TIME_SLOTS = [0, 120, 240, 360, 480, 600, 720]

    TASK_PRIORITY = {
        "装船": 1, "卸船": 1,
        "进箱": 2, "提箱": 2,
        "翻捣箱": 3
    }

    WEIGHTS = {
        'makespan': 0.5, 'balance': 0.3, 'time_balance': 0.05,
        'block_switch': 0.08, 'priority_delay': 0.05,
        'conflict': 0.15, 'idle': 0.02, 'track_congestion': 0.05
    }


# ================================
# 3. 数据加载类
# ================================
class YardDataLoader:
    """数据加载器"""

    def __init__(self, config: RealYardConfig):
        self.config = config

    def load_yard_layout(self):
        """加载堆场布局"""
        print("正在加载堆场布局...")

        yards, tracks = generate_yard_model(yard_coords, TRACK_WIDTH, TRACK_SPACING)

        # 处理analyze_track_distribution的返回值
        try:
            analysis_result = analyze_track_distribution(yards, tracks)
            if analysis_result is not None:
                area_stats, yard_track_details = analysis_result
            else:
                area_stats, yard_track_details = {}, {}
        except Exception as e:
            print(f"轨道分析失败: {e}")
            area_stats, yard_track_details = {}, {}

        # 转换为BlockInfo对象
        blocks = {}
        for yard in yards:
            block_id = yard.get('id', 'UNKNOWN')
            area = block_id[0] if block_id != 'UNKNOWN' else 'A'
            number_str = block_id[1:] if len(block_id) > 1 else '0'
            number = int(number_str) if number_str.isdigit() else 0
            has_power = (area == 'C' and number <= 20) or (area == 'B' and number <= 10)

            blocks[block_id] = BlockInfo(
                block_id=block_id,
                area=area,
                number=number,
                position_x=int(yard.get('center_x', 0)),
                position_y=int(yard.get('center_y', 0)),
                capacity=180,
                current_usage=random.randint(40, 150),
                has_power=has_power,
                track_connections=[],
                bay_count=35,
                row_count=6,
                tier_count=4
            )

        # 转换轨道数据
        track_objs = {}
        for i, track in enumerate(tracks):
            track_id = track.get('id', f'T{i + 1:03d}')  # 确保有valid的track_id
            track_objs[track_id] = TrackInfo(
                track_id=track_id,
                track_type=track.get('type', 'horizontal'),
                start_point=(track.get('x1', 0), track.get('y1', 0)),
                end_point=(track.get('x2', 0), track.get('y2', 0)),
                connected_blocks=[],
                track_width=track.get('width', TRACK_WIDTH),
                max_rtg_count=3 if track.get('type', 'horizontal') == 'horizontal' else 2,
                current_rtg_list=[]
            )

        # 建立箱区与轨道连接关系
        self._establish_track_connections(blocks, track_objs)

        print(f"堆场布局加载完成：{len(blocks)}个箱区，{len(track_objs)}条轨道")
        return blocks, track_objs

    def _establish_track_connections(self, blocks, tracks):
        """建立箱区与轨道的连接关系"""
        for block_id, block_info in blocks.items():
            connected_tracks = []

            for track_id, track_info in tracks.items():
                if self._is_track_accessible_to_block(block_info, track_info):
                    connected_tracks.append(track_id)
                    track_info.connected_blocks.append(block_id)

            block_info.track_connections = connected_tracks

            # 如果没有连接的轨道，强制连接最近的轨道
            if not connected_tracks:
                nearest_track = self._find_nearest_track(block_info, tracks)
                if nearest_track:
                    block_info.track_connections = [nearest_track]
                    tracks[nearest_track].connected_blocks.append(block_id)

    def _is_track_accessible_to_block(self, block_info, track_info):
        """判断轨道是否可被箱区访问"""
        block_x, block_y = block_info.position_x, block_info.position_y
        threshold = 120

        if track_info.track_type == 'horizontal':
            track_y = track_info.start_point[1]
            track_x_start = min(track_info.start_point[0], track_info.end_point[0])
            track_x_end = max(track_info.start_point[0], track_info.end_point[0])
            return (track_x_start - 30 <= block_x <= track_x_end + 30 and
                    abs(block_y - track_y) <= threshold)
        else:
            track_x = track_info.start_point[0]
            track_y_start = min(track_info.start_point[1], track_info.end_point[1])
            track_y_end = max(track_info.start_point[1], track_info.end_point[1])
            return (track_y_start - 30 <= block_y <= track_y_end + 30 and
                    abs(block_x - track_x) <= threshold)

    def _find_nearest_track(self, block_info, tracks):
        """找到距离箱区最近的轨道"""
        min_distance = float('inf')
        nearest_track_id = None

        for track_id, track_info in tracks.items():
            distance = self._calculate_distance_to_track(block_info, track_info)
            if distance < min_distance:
                min_distance = distance
                nearest_track_id = track_id

        return nearest_track_id

    def _calculate_distance_to_track(self, block_info, track_info):
        """计算箱区到轨道的距离"""
        block_x, block_y = block_info.position_x, block_info.position_y
        x1, y1 = track_info.start_point
        x2, y2 = track_info.end_point

        # 点到线段的距离
        A = block_x - x1
        B = block_y - y1
        C = x2 - x1
        D = y2 - y1

        dot = A * C + B * D
        len_sq = C * C + D * D

        if len_sq == 0:
            return (A * A + B * B) ** 0.5

        param = dot / len_sq

        if param < 0:
            xx, yy = x1, y1
        elif param > 1:
            xx, yy = x2, y2
        else:
            xx = x1 + param * C
            yy = y1 + param * D

        dx = block_x - xx
        dy = block_y - yy
        return (dx * dx + dy * dy) ** 0.5

    def load_tasks_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载任务数据"""
        if not load_workbook:
            return self._generate_sample_tasks()

        try:
            wb = load_workbook(filename)
            sheet = wb['Sheet1']
            tasks = []

            for row in sheet.iter_rows(min_row=2):
                if row[0].value is None:
                    continue

                tasks.append(TaskInfo(
                    task_id=str(row[0].value),
                    block_id=str(row[1].value),
                    bay=int(row[2].value),
                    row=int(row[3].value),
                    tier=int(row[4].value),
                    prep_time=str(row[5].value),
                    task_type=str(row[6].value),
                    truck_id=str(row[7].value) if row[7].value else None,
                    is_cold_container=row[8].value == '是' if row[8].value else False,
                    manual_mark=row[9].value == '是' if row[9].value else False,
                    estimated_duration=random.uniform(8, 25)
                ))

            print(f"从Excel加载了{len(tasks)}个任务")
            return tasks
        except Exception as e:
            print(f"加载Excel失败: {e}")
            return self._generate_sample_tasks()

    def _generate_sample_tasks(self):
        """生成示例任务数据"""
        sample_tasks = []
        task_types = ["装船", "卸船", "进箱", "提箱", "翻捣箱"]
        block_ids = ["A01", "A02", "B01", "B02", "C01", "C02", "D14", "D15"]

        for i in range(60):
            sample_tasks.append(TaskInfo(
                task_id=f"TASK{i + 1:03d}",
                block_id=random.choice(block_ids),
                bay=random.randint(1, 35),
                row=random.randint(1, 6),
                tier=random.randint(1, 4),
                prep_time="08:00",
                task_type=random.choice(task_types),
                truck_id=f"TR{random.randint(1, 20):03d}" if random.random() > 0.3 else None,
                is_cold_container=random.random() < 0.12,
                manual_mark=random.random() < 0.1,
                estimated_duration=random.uniform(8, 25)
            ))

        print(f"生成了{len(sample_tasks)}个示例任务")
        return sample_tasks

    def load_rtgs_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载RTG数据"""
        if not load_workbook:
            return self._generate_sample_rtgs()

        try:
            wb = load_workbook(filename)
            sheet = wb['Sheet2']
            rtgs = []

            track_ranges = {
                'T001-T066': list(range(1, 67)),
                'T067-T156': list(range(67, 157)),
                'T157-T270': list(range(157, 271)),
                'T271-T306': list(range(271, 307))
            }

            for row in sheet.iter_rows(min_row=2):
                if row[0].value is None:
                    continue

                rtg_name = str(row[0].value)
                rtg_id = int(rtg_name[3:]) if len(rtg_name) > 3 else random.randint(1, 100)
                rtg_tracks = str(row[1].value)
                rtg_type = str(row[2].value).lower()

                current_track_id = None
                for track_range, track_nums in track_ranges.items():
                    if rtg_tracks == track_range and track_nums:
                        track_num = random.choice(track_nums)
                        current_track_id = f'T{track_num:03d}'
                        break

                if not current_track_id:
                    current_track_id = f'T{random.randint(1, 306):03d}'

                available_areas = ['C', 'B'] if rtg_type == 'electric' else ['A', 'B', 'C', 'D']

                rtgs.append(RTGInfo(
                    rtg_id=rtg_id,
                    rtg_type=rtg_type,
                    current_track_id=current_track_id,
                    current_position=(random.randint(100, 1000), random.randint(100, 2000)),
                    login_status=str(row[3].value),
                    fault_status=str(row[4].value),
                    move_speed_horizontal=0.3,
                    move_speed_vertical=0.18,
                    operation_speed=2.8,
                    available_areas=available_areas,
                    max_lift_height=5
                ))

            print(f"从Excel加载了{len(rtgs)}台RTG")
            return rtgs
        except Exception as e:
            print(f"加载Excel RTG失败: {e}")
            return self._generate_sample_rtgs()

    def _generate_sample_rtgs(self):
        """生成示例RTG数据"""
        sample_rtgs = []
        for i in range(12):
            rtg_type = 'electric' if i < 4 else 'diesel'
            available_areas = ['C', 'B'] if rtg_type == 'electric' else ['A', 'B', 'C', 'D']

            sample_rtgs.append(RTGInfo(
                rtg_id=i + 1,
                rtg_type=rtg_type,
                current_track_id=f'T{random.randint(1, 306):03d}',
                current_position=(random.randint(100, 1000), random.randint(100, 2000)),
                login_status='online',
                fault_status='normal',
                move_speed_horizontal=0.3,
                move_speed_vertical=0.18,
                operation_speed=2.8,
                available_areas=available_areas,
                max_lift_height=5
            ))

        print(f"生成了{len(sample_rtgs)}台示例RTG")
        return sample_rtgs

    def load_trucks(self, tasks):
        """生成拖车数据"""
        trucks = []
        truck_tasks = {}

        for task in tasks:
            if task.truck_id:
                if task.truck_id not in truck_tasks:
                    truck_tasks[task.truck_id] = []
                truck_tasks[task.truck_id].append(task.task_id)

        for truck_id, task_ids in truck_tasks.items():
            ref_task = next(t for t in tasks if t.task_id == task_ids[0])
            trucks.append(TruckInfo(
                truck_id=truck_id,
                current_block_id=ref_task.block_id,
                current_bay=ref_task.bay,
                speed=random.uniform(15, 22),
                containers=task_ids[:2],
                arrival_time=random.uniform(0, 45)
            ))

        print(f"生成了{len(trucks)}台拖车")
        return trucks


# ================================
# 4. 可视化类
# ================================
class RealYardVisualization:
    """可视化类"""

    def __init__(self, config: RealYardConfig):
        self.config = config

    def plot_yard_layout(self, blocks: Dict[str, BlockInfo], tracks: Dict[str, TrackInfo],
                         save_path: str = 'real_yard_layout.png'):
        """绘制堆场布局图"""
        fig, ax = plt.subplots(figsize=(18, 14))

        self._draw_blocks(ax, blocks)
        self._draw_tracks(ax, tracks)
        self._add_area_labels(ax)

        ax.set_title('基于yard.py的集装箱堆场布局图', fontsize=16, fontweight='bold')
        ax.set_xlabel('X坐标 (堆场宽度方向)', fontsize=12)
        ax.set_ylabel('Y坐标 (堆场长度方向)', fontsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_aspect('equal')

        power_patch = mpatches.Patch(color='lightgreen', label='有电箱区')
        no_power_patch = mpatches.Patch(color='lightcoral', label='无电箱区')
        h_track_patch = mpatches.Patch(color='red', label='横向轨道')
        v_track_patch = mpatches.Patch(color='blue', label='竖向轨道')
        ax.legend(handles=[power_patch, no_power_patch, h_track_patch, v_track_patch],
                  loc='upper left', bbox_to_anchor=(0.02, 0.98))

        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"实际堆场布局图已保存到: {save_path}")
        plt.show()

    def _draw_blocks(self, ax, blocks):
        """绘制箱区"""
        for block_id, block_info in blocks.items():
            x, y = block_info.position_x, block_info.position_y
            color = 'lightgreen' if block_info.has_power else 'lightcoral'

            rect = plt.Rectangle((x - 30, y - 30), 60, 60,
                                 facecolor=color, edgecolor='black', linewidth=1.2)
            ax.add_patch(rect)
            ax.text(x, y, block_id, ha='center', va='center', fontsize=8, fontweight='bold')

            usage_rate = block_info.current_usage / block_info.capacity
            ax.text(x, y - 15, f"{usage_rate:.1%}", ha='center', va='center',
                    fontsize=6, color='darkblue')

    def _draw_tracks(self, ax, tracks):
        """绘制轨道"""
        for track_id, track_info in tracks.items():
            start_x, start_y = track_info.start_point
            end_x, end_y = track_info.end_point

            color = 'red' if track_info.track_type == 'horizontal' else 'blue'
            ax.plot([start_x, end_x], [start_y, end_y], color=color, linewidth=2, alpha=0.8)

    def _add_area_labels(self, ax):
        """添加区域标签"""
        area_centers = {'A': (900, 500), 'B': (650, 800), 'C': (380, 1000), 'D': (180, 500)}
        for area, (x, y) in area_centers.items():
            ax.text(x, y, f"{area}区", ha='center', va='center',
                    fontsize=14, fontweight='bold', color='darkblue',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="yellow", alpha=0.7))


# ================================
# 5. RTG调度算法
# ================================
class AdvancedRTGScheduler:
    """高级RTG调度算法"""

    def __init__(self, config: RealYardConfig):
        self.config = config
        self.tasks = []
        self.rtgs = []
        self.trucks = []
        self.blocks = {}
        self.tracks = {}
        self.BOX_LAYOUT = {}

    def initialize_data(self, tasks, rtgs, trucks, blocks, tracks):
        """初始化调度数据"""
        self.tasks = tasks
        self.rtgs = rtgs
        self.trucks = trucks
        self.blocks = blocks
        self.tracks = tracks
        self._initialize_box_layout()
        self._validate_rtg_assignments()

    def _initialize_box_layout(self):
        """初始化箱区布局"""
        area_blocks = {'A': [], 'B': [], 'C': [], 'D': []}

        for block_id, block_info in self.blocks.items():
            area_blocks[block_info.area].append((block_id, block_info))

        for area, blocks_list in area_blocks.items():
            blocks_list.sort(key=lambda x: x[1].number)
            area_base_col = {'A': 0, 'B': 10, 'C': 20, 'D': 30}[area]

            for i, (block_id, block_info) in enumerate(blocks_list):
                if area == 'A':
                    row, col = i // 4, area_base_col + (i % 4)
                elif area in ['B', 'C']:
                    row, col = i // 4, area_base_col + (i % 4)
                else:  # D区
                    row, col = i // 3, area_base_col + (i % 3)

                self.BOX_LAYOUT[block_id] = {'row': row, 'col': col, 'area': area}

    def _validate_rtg_assignments(self):
        """验证RTG轨道分配"""
        valid_tracks = set(self.tracks.keys())

        for rtg in self.rtgs:
            if rtg.current_track_id not in valid_tracks:
                # 找到适合该RTG的轨道
                suitable_tracks = []
                for track_id, track_info in self.tracks.items():
                    for block_id in track_info.connected_blocks:
                        if block_id in self.blocks:
                            block_area = self.blocks[block_id].area
                            if block_area in rtg.available_areas:
                                suitable_tracks.append(track_id)
                                break

                if suitable_tracks:
                    rtg.current_track_id = suitable_tracks[0]
                    print(f"为RTG-{rtg.rtg_id}重新分配轨道: {rtg.current_track_id}")
                else:
                    # 分配第一个可用轨道
                    if self.tracks:
                        rtg.current_track_id = list(self.tracks.keys())[0]
                        print(f"为RTG-{rtg.rtg_id}分配默认轨道: {rtg.current_track_id}")
                    else:
                        rtg.current_track_id = 'T001'
                        print(f"为RTG-{rtg.rtg_id}分配虚拟轨道: T001")

    def calculate_block_distance(self, block1, block2):
        """计算箱区距离"""
        if block1 == block2:
            return 0, 0

        if block1 not in self.BOX_LAYOUT or block2 not in self.BOX_LAYOUT:
            return 0, 0

        row1, col1 = self.BOX_LAYOUT[block1]["row"], self.BOX_LAYOUT[block1]["col"]
        row2, col2 = self.BOX_LAYOUT[block2]["row"], self.BOX_LAYOUT[block2]["col"]

        return abs(row1 - row2), abs(col1 - col2)

    def calculate_move_time(self, current_pos, target_pos, block_current, block_target):
        """计算移动时间"""
        block_row_distance, block_col_distance = self.calculate_block_distance(block_current, block_target)

        if block_row_distance == 0:
            horizontal_move_time = abs(current_pos - target_pos) * 0.28
        else:
            if current_pos < 10 and target_pos < 10:
                horizontal_move_time = (current_pos + target_pos) * 0.28 + 3
            elif current_pos > 10 and target_pos > 10:
                horizontal_move_time = abs(current_pos - 60) * 0.28 + abs(target_pos - 60) * 0.28 + 3
            elif current_pos > 10 and target_pos < 10:
                horizontal_move_time = ((19 - current_pos) + target_pos) * 0.28 + 3
            else:
                horizontal_move_time = (abs(19 - target_pos) + current_pos) * 0.28 + 3

        if block_col_distance == 0 or block_col_distance == 1:
            vertical_move_time = block_col_distance * 10
        else:
            vertical_move_time = 13

        return horizontal_move_time + vertical_move_time + 3

    def assign_rtg_to_task(self, task: TaskInfo) -> Optional[RTGInfo]:
        """为任务分配RTG"""
        if task.block_id not in self.blocks:
            return None

        block_info = self.blocks[task.block_id]
        available_rtgs = []

        # 筛选可用RTG
        for rtg in self.rtgs:
            if (rtg.login_status == 'online' and
                    rtg.fault_status == 'normal' and
                    block_info.area in rtg.available_areas):

                # 检查RTG是否能到达该箱区
                if self._can_rtg_reach_block(rtg, block_info):
                    available_rtgs.append(rtg)

        if not available_rtgs:
            return None

        # 冷箱优先使用电动RTG
        if task.is_cold_container and block_info.has_power:
            electric_rtgs = [rtg for rtg in available_rtgs if rtg.rtg_type == 'electric']
            if electric_rtgs:
                available_rtgs = electric_rtgs

        # 选择时间成本最小的RTG
        best_rtg = min(available_rtgs, key=lambda r: self._calculate_rtg_cost(r, task))
        return best_rtg

    def _can_rtg_reach_block(self, rtg: RTGInfo, block_info: BlockInfo) -> bool:
        """判断RTG是否能到达箱区"""
        # 检查轨道连通性
        if not block_info.track_connections:
            return True  # 如果箱区没有轨道连接，允许RTG到达

        # 检查RTG当前轨道是否能到达目标箱区的轨道
        rtg_track = self.tracks.get(rtg.current_track_id)
        if not rtg_track:
            return True

        # 简化连通性判断：同区域或相邻区域认为可达
        for target_track_id in block_info.track_connections:
            target_track = self.tracks.get(target_track_id)
            if target_track and self._tracks_connected(rtg_track, target_track):
                return True

        return True  # 默认认为可达

    def _tracks_connected(self, track1: TrackInfo, track2: TrackInfo) -> bool:
        """判断轨道是否连通"""
        if track1.track_id == track2.track_id:
            return True

        # 检查是否有共同连接的箱区
        common_blocks = set(track1.connected_blocks) & set(track2.connected_blocks)
        if common_blocks:
            return True

        # 检查距离
        points1 = [track1.start_point, track1.end_point]
        points2 = [track2.start_point, track2.end_point]

        for p1 in points1:
            for p2 in points2:
                distance = ((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2) ** 0.5
                if distance <= 100:
                    return True

        return False

    def _calculate_rtg_cost(self, rtg: RTGInfo, task: TaskInfo) -> float:
        """计算RTG执行任务的成本"""
        # 获取RTG当前位置对应的箱区
        rtg_block = self._get_rtg_current_block(rtg)
        if not rtg_block:
            return float('inf')

        # 估算当前贝位
        rtg_bay = self._estimate_rtg_bay(rtg)
        target_bay = task.bay

        # 计算移动时间
        move_time = self.calculate_move_time(rtg_bay, target_bay, rtg_block, task.block_id)

        return move_time + task.estimated_duration

    def _get_rtg_current_block(self, rtg: RTGInfo) -> Optional[str]:
        """获取RTG当前所在箱区"""
        if rtg.current_track_id in self.tracks:
            track = self.tracks[rtg.current_track_id]
            if track.connected_blocks:
                return track.connected_blocks[0]

        # 根据坐标找最近箱区
        rtg_x, rtg_y = rtg.current_position
        min_distance = float('inf')
        closest_block = None

        for block_id, block_info in self.blocks.items():
            if block_info.area in rtg.available_areas:
                distance = ((rtg_x - block_info.position_x) ** 2 +
                            (rtg_y - block_info.position_y) ** 2) ** 0.5
                if distance < min_distance:
                    min_distance = distance
                    closest_block = block_id

        return closest_block

    def _estimate_rtg_bay(self, rtg: RTGInfo) -> int:
        """估算RTG当前贝位"""
        return random.randint(1, 35)  # 简化处理

    def preprocess_tasks(self):
        """任务预处理"""
        manual_tasks = [task for task in self.tasks if task.manual_mark]
        normal_tasks = [task for task in self.tasks if not task.manual_mark]

        manual_tasks.sort(key=lambda t: (0 if t.is_cold_container else 1))

        def task_sort_key(task):
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            cold_priority = 0 if task.is_cold_container else 1
            return (task_priority, cold_priority)

        normal_tasks.sort(key=task_sort_key)
        return manual_tasks, normal_tasks

    def optimize_schedule(self):
        """优化调度方案"""
        print("\n正在优化RTG调度方案...")
        manual_tasks, normal_tasks = self.preprocess_tasks()

        schedule = {
            'manual_task_assignments': [],
            'normal_task_assignments': [],
            'total_makespan': 0,
            'rtg_utilization': {},
            'priority_summary': {}
        }

        rtg_workload = {rtg.rtg_id: 0 for rtg in self.rtgs}
        total_tasks = len(self.tasks)
        assigned_count = 0

        # 处理人工标记任务
        print(f"处理人工标记任务 ({len(manual_tasks)}个):")
        for i, task in enumerate(manual_tasks):
            rtg = self.assign_rtg_to_task(task)
            if rtg:
                start_time = rtg_workload[rtg.rtg_id]
                assignment = {
                    'task_id': task.task_id,
                    'rtg_id': rtg.rtg_id,
                    'block_id': task.block_id,
                    'task_type': task.task_type,
                    'estimated_time': task.estimated_duration,
                    'is_cold': task.is_cold_container,
                    'priority_level': 1,
                    'start_time': start_time,
                    'end_time': start_time + task.estimated_duration
                }
                schedule['manual_task_assignments'].append(assignment)
                rtg_workload[rtg.rtg_id] += task.estimated_duration
                assigned_count += 1

                cold_mark = "(冷箱)" if task.is_cold_container else ""
                print(f"  {i + 1}. {task.task_id}: {task.task_type}{cold_mark} -> RTG-{rtg.rtg_id}")

        # 处理普通任务
        print(f"处理普通任务 ({len(normal_tasks)}个):")
        current_priority = None
        task_count = 0

        for task in normal_tasks:
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            if current_priority != task_priority:
                if current_priority is not None:
                    print(f"  优先级{current_priority}完成，共{task_count}个任务")
                current_priority = task_priority
                task_count = 0
                priority_names = {1: "装卸类", 2: "收发类", 3: "翻倒类"}
                print(f"  开始处理优先级{task_priority}({priority_names.get(task_priority, '其他')}):")

            rtg = self.assign_rtg_to_task(task)
            if rtg:
                start_time = rtg_workload[rtg.rtg_id]
                assignment = {
                    'task_id': task.task_id,
                    'rtg_id': rtg.rtg_id,
                    'block_id': task.block_id,
                    'task_type': task.task_type,
                    'estimated_time': task.estimated_duration,
                    'is_cold': task.is_cold_container,
                    'priority_level': task_priority + 1,
                    'start_time': start_time,
                    'end_time': start_time + task.estimated_duration
                }
                schedule['normal_task_assignments'].append(assignment)
                rtg_workload[rtg.rtg_id] += task.estimated_duration
                assigned_count += 1
                task_count += 1

                cold_mark = "(冷箱)" if task.is_cold_container else ""
                print(f"    {task.task_id}: {task.task_type}{cold_mark} -> RTG-{rtg.rtg_id}")

        if current_priority is not None:
            print(f"  优先级{current_priority}完成，共{task_count}个任务")

        # 计算结果
        schedule['rtg_utilization'] = rtg_workload
        schedule['total_makespan'] = max(rtg_workload.values()) if rtg_workload.values() else 0

        # 统计优先级分布
        priority_stats = {}
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
        for assignment in all_assignments:
            priority = assignment['priority_level']
            task_type = assignment['task_type']
            is_cold = assignment['is_cold']

            if priority not in priority_stats:
                priority_stats[priority] = {'total': 0, 'cold': 0, 'types': {}}

            priority_stats[priority]['total'] += 1
            if is_cold:
                priority_stats[priority]['cold'] += 1
            if task_type not in priority_stats[priority]['types']:
                priority_stats[priority]['types'][task_type] = 0
            priority_stats[priority]['types'][task_type] += 1

        schedule['priority_summary'] = priority_stats

        print(f"\n调度优化完成:")
        print(f"  - 总任务数: {total_tasks}")
        print(f"  - 成功分配: {assigned_count}个 ({assigned_count / total_tasks * 100:.1f}%)")
        print(f"  - 预计总时间: {schedule['total_makespan']:.1f}分钟")

        if schedule['total_makespan'] > 0:
            avg_utilization = sum(rtg_workload.values()) / len(rtg_workload) / schedule['total_makespan'] * 100
            print(f"  - 平均RTG利用率: {avg_utilization:.1f}%")

        return schedule

    def export_schedule_to_excel(self, schedule, filename=None):
        """导出调度结果到Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"RTG调度结果_{timestamp}.xlsx"

        print(f"正在导出调度结果到: {filename}")

        wb = Workbook()
        wb.remove(wb.active)

        # 详细调度结果
        ws_detail = wb.create_sheet("详细调度结果")
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']

        headers = ['任务ID', '箱区', '贝位', '排', '层', '任务类型', '分配RTG',
                   '开始时间(分钟)', '结束时间(分钟)', '作业时长(分钟)', '是否冷箱', '优先级']
        ws_detail.append(headers)

        for assignment in all_assignments:
            task_detail = next((t for t in self.tasks if t.task_id == assignment['task_id']), None)
            row = [
                assignment['task_id'],
                assignment['block_id'],
                task_detail.bay if task_detail else '',
                task_detail.row if task_detail else '',
                task_detail.tier if task_detail else '',
                assignment['task_type'],
                f"RTG-{assignment['rtg_id']}",
                round(assignment['start_time'], 2),
                round(assignment['end_time'], 2),
                round(assignment['estimated_time'], 2),
                '是' if assignment['is_cold'] else '否',
                assignment['priority_level']
            ]
            ws_detail.append(row)

        # RTG工作负载统计
        ws_rtg = wb.create_sheet("RTG工作负载")
        ws_rtg.append(['RTG编号', 'RTG类型', '任务数量', '总工作时间', '利用率(%)'])

        rtg_stats = {}
        for assignment in all_assignments:
            rtg_id = assignment['rtg_id']
            if rtg_id not in rtg_stats:
                rtg_info = next((r for r in self.rtgs if r.rtg_id == rtg_id), None)
                rtg_stats[rtg_id] = {
                    'type': rtg_info.rtg_type if rtg_info else 'unknown',
                    'tasks': 0,
                    'time': 0
                }
            rtg_stats[rtg_id]['tasks'] += 1
            rtg_stats[rtg_id]['time'] += assignment['estimated_time']

        total_makespan = schedule['total_makespan']
        for rtg_id, stats in rtg_stats.items():
            utilization = stats['time'] / total_makespan * 100 if total_makespan > 0 else 0
            ws_rtg.append([
                f"RTG-{rtg_id}",
                stats['type'],
                stats['tasks'],
                round(stats['time'], 2),
                round(utilization, 2)
            ])

        wb.save(filename)
        print(f"调度结果已导出到: {filename}")
        return filename


# ================================
# 6. 主函数
# ================================
def main():
    """主函数"""
    print("=== 基于实际堆场的RTG调度系统初始化 ===")

    config = RealYardConfig()
    data_loader = YardDataLoader(config)

    # 加载数据
    blocks, tracks = data_loader.load_yard_layout()
    tasks = data_loader.load_tasks_from_excel()
    rtgs = data_loader.load_rtgs_from_excel()
    trucks = data_loader.load_trucks(tasks)

    print(f"基于实际堆场的数据加载完成:")
    print(f"  - 任务数量: {len(tasks)}")
    print(f"  - RTG数量: {len(rtgs)}")
    print(f"  - 拖车数量: {len(trucks)}")
    print(f"  - 箱区数量: {len(blocks)}")
    print(f"  - 轨道数量: {len(tracks)}")

    # 统计信息
    area_stats = {}
    power_stats = {'有电': 0, '无电': 0}
    for block_info in blocks.values():
        area = block_info.area
        area_stats[area] = area_stats.get(area, 0) + 1
        if block_info.has_power:
            power_stats['有电'] += 1
        else:
            power_stats['无电'] += 1

    print(f"=== 堆场统计信息 ===")
    print("各区域箱区分布:")
    for area in sorted(area_stats.keys()):
        print(f"  {area}区: {area_stats[area]}个箱区")
    print(f"电力设施分布:")
    for power_type, count in power_stats.items():
        print(f"  {power_type}箱区: {count}个")

    # 可视化
    visualizer = RealYardVisualization(config)
    visualizer.plot_yard_layout(blocks, tracks)

    # 显示详细信息
    print("=== 箱区信息样例 ===")
    for i, (block_id, block_info) in enumerate(list(blocks.items())[:8]):
        power_status = "有电" if block_info.has_power else "无电"
        usage_rate = block_info.current_usage / block_info.capacity * 100
        track_count = len(block_info.track_connections)
        print(f"{block_id}: {block_info.area}区, 位置({block_info.position_x},{block_info.position_y}), "
              f"{power_status}, 使用率{usage_rate:.1f}%, 连接{track_count}条轨道")

    print("=== 轨道信息样例 ===")
    for i, (track_id, track_info) in enumerate(list(tracks.items())[:6]):
        track_type_name = "横向轨道" if track_info.track_type == 'horizontal' else "竖向轨道"
        block_count = len(track_info.connected_blocks)
        rtg_count = len(track_info.current_rtg_list)
        print(f"{track_id}: {track_type_name}, 连接{block_count}个箱区, 当前{rtg_count}台RTG")

    # RTG配置信息
    electric_count = sum(1 for rtg in rtgs if rtg.rtg_type == 'electric')
    diesel_count = len(rtgs) - electric_count
    print("=== RTG配置信息 ===")
    print(f"电动RTG: {electric_count}台 (主要服务有电箱区)")
    print(f"柴油RTG: {diesel_count}台 (服务所有箱区)")

    print("RTG轨道分布:")
    rtg_track_distribution = {}
    for rtg in rtgs[:5]:  # 显示前5台RTG
        track_id = rtg.current_track_id
        if track_id not in rtg_track_distribution:
            rtg_track_distribution[track_id] = []
        rtg_track_distribution[track_id].append(rtg.rtg_id)

    for track_id, rtg_list in rtg_track_distribution.items():
        print(f"  {track_id}: RTG {rtg_list}")

    # 任务分布信息
    task_area_stats = {}
    cold_count = 0
    task_type_stats = {}

    for task in tasks:
        area = task.block_id[0] if task.block_id else 'Unknown'
        task_area_stats[area] = task_area_stats.get(area, 0) + 1

        if task.is_cold_container:
            cold_count += 1

        task_type_stats[task.task_type] = task_type_stats.get(task.task_type, 0) + 1

    print("=== 任务分布信息 ===")
    print("各区域任务分布:")
    for area in sorted(task_area_stats.keys()):
        print(f"  {area}区: {task_area_stats[area]}个任务")

    print(f"冷箱任务: {cold_count}个 ({cold_count / len(tasks) * 100:.1f}%)")
    print("任务类型分布:")
    for task_type, count in task_type_stats.items():
        print(f"  {task_type}: {count}个")

    print("=== 系统优化目标 ===")
    objectives = [
        "1. 最小化总完成时间 (makespan)",
        "2. 平衡RTG工作负载",
        "3. 减少RTG跨区域移动",
        "4. 优先处理高优先级任务",
        "5. 避免轨道拥堵",
        "6. 提高冷箱任务处理效率"
    ]
    for obj in objectives:
        print(obj)

    print("=" * 50)
    print("运行高级RTG调度算法...")

    # 运行调度算法
    scheduler = AdvancedRTGScheduler(config)
    scheduler.initialize_data(tasks, rtgs, trucks, blocks, tracks)
    schedule = scheduler.optimize_schedule()

    # 显示结果
    print(f"\n=== 调度结果摘要 ===")
    print(f"人工标记任务: {len(schedule['manual_task_assignments'])}个")
    print(f"普通任务分配: {len(schedule['normal_task_assignments'])}个")
    print(f"预计完成时间: {schedule['total_makespan']:.1f}分钟")

    print(f"\n=== 优先级处理统计 ===")
    priority_names = {1: "人工标记", 2: "装卸类", 3: "收发类", 4: "翻倒类"}
    for priority, stats in schedule['priority_summary'].items():
        level_name = priority_names.get(priority, f"优先级{priority}")
        print(f"  {level_name}: {stats['total']}个任务 (冷箱{stats['cold']}个)")
        for task_type, count in stats['types'].items():
            print(f"    - {task_type}: {count}个")

    print(f"\n=== RTG工作负载分布 ===")
    rtg_items = sorted(schedule['rtg_utilization'].items(), key=lambda x: x[1], reverse=True)
    for rtg_id, workload in rtg_items[:8]:
        rtg_info = next((r for r in rtgs if r.rtg_id == rtg_id), None)
        if rtg_info and workload > 0:
            utilization = workload / schedule['total_makespan'] * 100
            print(f"  RTG-{rtg_id}({rtg_info.rtg_type}): {workload:.1f}分钟 (利用率{utilization:.1f}%)")

    # 冷箱处理统计
    all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
    cold_assignments = [a for a in all_assignments if a['is_cold']]

    print(f"\n=== 冷箱处理统计 ===")
    print(f"总冷箱任务: {cold_count}个")
    print(f"  - 人工标记冷箱: {sum(1 for a in schedule['manual_task_assignments'] if a['is_cold'])}个")
    print(f"  - 普通冷箱: {sum(1 for a in schedule['normal_task_assignments'] if a['is_cold'])}个")

    if cold_assignments:
        cold_task_types = {}
        for assignment in cold_assignments:
            task_type = assignment['task_type']
            cold_task_types[task_type] = cold_task_types.get(task_type, 0) + 1

        print(f"  冷箱任务类型分布:")
        for task_type, count in cold_task_types.items():
            print(f"    - {task_type}: {count}个")

    print(f"\n系统已完成基于实际堆场布局的RTG调度优化！")

    # 导出Excel
    try:
        excel_file = scheduler.export_schedule_to_excel(schedule)
        print(f"调度结果已导出到Excel: {excel_file}")
    except Exception as e:
        print(f"Excel导出失败: {e}")

    return tasks, rtgs, trucks, blocks, tracks, schedule


if __name__ == "__main__":
    tasks, rtgs, trucks, blocks, tracks, schedule = main()