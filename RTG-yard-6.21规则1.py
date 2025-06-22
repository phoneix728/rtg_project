import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import warnings
from openpyxl import Workbook
from datetime import datetime
import re
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Set
import heapq
from openpyxl import load_workbook

# 导入yard.py的功能
from yard import generate_yard_model, yard_coords, analyze_track_distribution, TRACK_WIDTH, TRACK_SPACING
# 设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
except:
    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
warnings.filterwarnings('ignore', category=UserWarning)

# ================================
# 1. 数据结构定义
# ================================
@dataclass
class BlockInfo:
    """箱区信息数据类"""
    block_id: str
    area: str
    number: int
    position_x: int
    position_y: int
    capacity: int
    current_usage: int
    has_power: bool
    track_connections: List[str]
    bay_count: int
    row_count: int
    tier_count: int


@dataclass
class TrackInfo:
    """轨道信息数据类"""
    track_id: str
    track_type: str
    start_point: Tuple[int, int]
    end_point: Tuple[int, int]
    connected_blocks: List[str]
    track_width: float
    max_rtg_count: int
    current_rtg_list: List[int]


@dataclass
class TaskInfo:
    """任务信息数据类"""
    task_id: str
    block_id: str
    bay: int
    row: int
    tier: int
    prep_time: str
    task_type: str
    truck_id: Optional[str] = None
    is_cold_container: bool = False
    manual_mark: bool = False
    estimated_duration: float = 0.0


@dataclass
class RTGInfo:
    """轮胎吊信息数据类"""
    rtg_id: int
    rtg_type: str
    current_track_id: str
    current_position: Tuple[int, int]
    login_status: str
    fault_status: str
    move_speed_horizontal: float
    move_speed_vertical: float
    operation_speed: float
    available_areas: List[str]
    max_lift_height: int
    current_load: Optional[str] = None

@dataclass
class TruckInfo:
    """拖车信息数据类"""
    truck_id: str
    current_block_id: str
    current_bay: int
    speed: float
    containers: List[str]
    arrival_time: float
    route_plan: List[str] = None


@dataclass
class RTGTaskQueue:
    """RTG任务队列"""
    rtg_id: int
    tasks: List[Dict] = field(default_factory=list)  # 任务队列
    current_position: Tuple[int, int] = (0, 0)  # 当前位置
    current_block: str = ""  # 当前箱区
    current_bay: int = 0  # 当前贝位
    finish_time: float = 0.0  # 当前所有任务完成时间

    def add_task(self, task_assignment: Dict):
        """添加任务到队列"""
        self.tasks.append(task_assignment)
        self.finish_time = task_assignment['end_time']
        self.current_block = task_assignment['block_id']
        self.current_bay = task_assignment['bay']

    def get_next_available_time(self) -> float:
        """获取RTG下次可用时间"""
        return self.finish_time

    def get_current_location(self) -> Tuple[str, int]:
        """获取RTG当前位置（箱区，贝位）"""
        if not self.tasks:
            # 如果没有任务，返回RTG初始位置
            return self.current_block, self.current_bay
        else:
            # 返回最后一个任务的位置
            last_task = self.tasks[-1]
            return last_task['block_id'], last_task['bay']


# ================================
# 2. 系统配置类
# ================================
class RealYardConfig:
    """系统配置"""
    POPULATION_SIZE = 150
    MAX_GENERATIONS = 300
    CROSSOVER_RATE = 0.9
    MUTATION_RATE = 0.25
    ELITE_RATE = 0.1
    NUM_CRANES = 12
    TIME_SLOTS = [0, 120, 240, 360, 480, 600, 720]

    TASK_PRIORITY = {
        "装船": 1, "卸船": 1,
        "进箱": 2, "提箱": 2,
        "翻捣箱": 3
    }

    WEIGHTS = {
        'makespan': 0.5, 'balance': 0.3, 'time_balance': 0.05,
        'block_switch': 0.08, 'priority_delay': 0.05,
        'conflict': 0.15, 'idle': 0.02, 'track_congestion': 0.05
    }


# ================================
# 3. 数据加载类
# ================================
class YardDataLoader:
    """数据加载器 """
    def __init__(self, config: RealYardConfig):
        self.config = config

    def load_yard_layout(self):
        """加载堆场布局"""
        print("正在加载堆场布局...")
        yards, tracks = generate_yard_model(yard_coords, TRACK_WIDTH, TRACK_SPACING)
        # 处理analyze_track_distribution的返回值
        try:
            analysis_result = analyze_track_distribution(yards, tracks)
            if analysis_result is not None:
                area_stats, yard_track_details = analysis_result
            else:
                area_stats, yard_track_details = {}, {}
        except Exception as e:
            print(f"轨道分析失败: {e}")
            area_stats, yard_track_details = {}, {}

        # 转换为BlockInfo对象
        blocks = {}
        for yard in yards:
            block_id = yard.get('id', 'UNKNOWN')
            area = block_id[0] if block_id != 'UNKNOWN' else 'A'
            number_str = block_id[1:] if len(block_id) > 1 else '0'
            number = int(number_str) if number_str.isdigit() else 0
            has_power = (area == 'C' and number <= 20) or (area == 'B' and number <= 10)

            blocks[block_id] = BlockInfo(
                block_id=block_id,
                area=area,
                number=number,
                position_x=int(yard.get('center_x', 0)),
                position_y=int(yard.get('center_y', 0)),
                capacity=180,
                current_usage=random.randint(40, 150),
                has_power=has_power,
                track_connections=[],  # 稍后建立连接关系
                bay_count=19,
                row_count=6,
                tier_count=6
            )

        # 转换轨道数据并建立ID映射
        track_objs = {}
        track_id_to_name = {}  # track_id到track_name的映射
        track_name_to_id = {}  # track_name到track_id的映射

        for i, track in enumerate(tracks):
            track_id = track.get('id', f'T{i + 1:03d}')
            track_name = track.get('name', track_id)  # 从yard.py获取的轨道名称

            track_objs[track_id] = TrackInfo(
                track_id=track_id,
                track_type=track.get('type', 'horizontal'),
                start_point=(track.get('x1', 0), track.get('y1', 0)),
                end_point=(track.get('x2', 0), track.get('y2', 0)),
                connected_blocks=[],
                track_width=track.get('width', TRACK_WIDTH),
                max_rtg_count=3 if track.get('type', 'horizontal') == 'horizontal' else 2,
                current_rtg_list=[]
            )

            # 建立ID和名称的双向映射
            track_id_to_name[track_id] = track_name
            track_name_to_id[track_name] = track_id

        # 建立箱区与轨道连接关系（新逻辑）
        self._establish_block_track_connections(blocks, track_objs, track_name_to_id)

        print(f"堆场布局加载完成：{len(blocks)}个箱区，{len(track_objs)}条轨道")
        return blocks, track_objs, track_id_to_name, track_name_to_id

    def _establish_block_track_connections(self, blocks, tracks, track_name_to_id):
        """
        建立箱区与轨道的连接关系
        直接从yard.py读取的轨道name中解析箱区ID
        """
        # 初始化所有箱区的轨道连接列表
        for block_id in blocks.keys():
            blocks[block_id].track_connections = []
        # 遍历所有轨道，根据name建立连接
        for track_name, track_id in track_name_to_id.items():
            # 从轨道名称中提取箱区ID（例如：A01_L1 -> A01）
            if '_' in track_name:
                block_id = track_name.split('_')[0]  # 取下划线前的部分作为箱区ID
                # 如果该箱区存在，建立连接关系
                if block_id in blocks:
                    blocks[block_id].track_connections.append(track_id)
                    tracks[track_id].connected_blocks.append(block_id)
                else:
                    print(f"警告：轨道{track_name}对应的箱区{block_id}不存在")

    def _parse_rtg_track_range(self, rtg_tracks_str):
        """
        解析RTG轨道范围字符串
        例如：'T001-T066' -> ['T001', 'T002', ..., 'T066']
        """
        if not rtg_tracks_str or rtg_tracks_str == 'None':
            return []
        try:
            # 匹配 T001-T066 格式
            match = re.match(r'T(\d+)-T(\d+)', rtg_tracks_str)
            if match:
                start_num = int(match.group(1))
                end_num = int(match.group(2))
                return [f'T{i:03d}' for i in range(start_num, end_num + 1)]
            else:
                # 如果不是范围格式，尝试作为单个轨道ID处理
                return [rtg_tracks_str] if rtg_tracks_str.startswith('T') else []
        except Exception as e:
            print(f"解析RTG轨道范围失败: {rtg_tracks_str}, 错误: {e}")
            return []


    def _calculate_distance_to_track(self, block_info, track_info):
        """计算箱区到轨道的距离"""
        block_x, block_y = block_info.position_x, block_info.position_y
        x1, y1 = track_info.start_point
        x2, y2 = track_info.end_point

        # 点到线段的距离
        A = block_x - x1
        B = block_y - y1
        C = x2 - x1
        D = y2 - y1

        dot = A * C + B * D
        len_sq = C * C + D * D

        if len_sq == 0:
            return (A * A + B * B) ** 0.5

        param = dot / len_sq

        if param < 0:
            xx, yy = x1, y1
        elif param > 1:
            xx, yy = x2, y2
        else:
            xx = x1 + param * C
            yy = y1 + param * D

        dx = block_x - xx
        dy = block_y - yy
        return (dx * dx + dy * dy) ** 0.5

    def load_tasks_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载任务数据"""
        wb = load_workbook(filename)
        sheet = wb['Sheet1']
        tasks = []

        for row in sheet.iter_rows(min_row=2):
            if row[0].value is None:
                continue

            tasks.append(TaskInfo(
                task_id=str(row[0].value),
                block_id=str(row[1].value),
                bay=int(row[2].value),
                row=int(row[3].value),
                tier=int(row[4].value),
                prep_time=str(row[5].value),
                task_type=str(row[6].value),
                truck_id=str(row[7].value) if row[7].value else None,
                is_cold_container=row[8].value == '是' if row[8].value else False,
                manual_mark=row[9].value == '是' if row[9].value else False,
                estimated_duration=random.uniform(8, 25)
            ))
        print(f"从Excel加载了{len(tasks)}个任务")
        return tasks

    def load_rtgs_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载RTG数据（修改版）"""
        wb = load_workbook(filename)
        sheet = wb['Sheet2']
        rtgs = []

        for row in sheet.iter_rows(min_row=2):
            if row[0].value is None:
                continue

            rtg_name = str(row[0].value)
            rtg_id = int(rtg_name[3:]) if len(rtg_name) > 3 else random.randint(1, 100)
            rtg_tracks_str = str(row[1].value)  # 例如：'T001-T066'
            rtg_type = str(row[2].value).lower()

            # 解析RTG可操作的轨道范围
            rtg_available_tracks = self._parse_rtg_track_range(rtg_tracks_str)
            # 随机选择一个当前轨道（在其范围内）
            current_track_id = rtg_available_tracks[0] if rtg_available_tracks else 'T001'
            available_areas = ['C', 'B'] if rtg_type == 'electric' else ['A', 'B', 'C', 'D']

            rtg_info = RTGInfo(
                rtg_id=rtg_id,
                rtg_type=rtg_type,
                current_track_id=current_track_id,
                current_position=(random.randint(100, 1000), random.randint(100, 2000)),
                login_status=str(row[3].value),
                fault_status=str(row[4].value),
                move_speed_horizontal=0.3,
                move_speed_vertical=0.18,
                operation_speed=2.8,
                available_areas=available_areas,
                max_lift_height=6
            )

            # 添加RTG可操作轨道范围属性
            rtg_info.available_tracks = rtg_available_tracks

            rtgs.append(rtg_info)

        print(f"从Excel加载了{len(rtgs)}台RTG")
        return rtgs

    def load_trucks(self, tasks):
        """生成拖车数据"""
        trucks = []
        truck_tasks = {}

        for task in tasks:
            if task.truck_id:
                if task.truck_id not in truck_tasks:
                    truck_tasks[task.truck_id] = []
                truck_tasks[task.truck_id].append(task.task_id)

        for truck_id, task_ids in truck_tasks.items():
            ref_task = next(t for t in tasks if t.task_id == task_ids[0])
            trucks.append(TruckInfo(
                truck_id=truck_id,
                current_block_id=ref_task.block_id,
                current_bay=ref_task.bay,
                speed=random.uniform(15, 22),
                containers=task_ids[:2],
                arrival_time=random.uniform(0, 45)
            ))
        print(f"生成了{len(trucks)}台拖车")
        return trucks


# ================================
# 4. 可视化类
# ================================
class RealYardVisualization:
    """可视化类"""
    def __init__(self, config: RealYardConfig):
        self.config = config
    def plot_yard_layout(self, blocks: Dict[str, BlockInfo], tracks: Dict[str, TrackInfo],
                         save_path: str = 'real_yard_layout.png'):
        """绘制堆场布局图"""
        fig, ax = plt.subplots(figsize=(18, 14))

        self._draw_blocks(ax, blocks)
        self._draw_tracks(ax, tracks)
        self._add_area_labels(ax)

        ax.set_title('基于yard.py的集装箱堆场布局图', fontsize=16, fontweight='bold')
        ax.set_xlabel('X坐标 (堆场宽度方向)', fontsize=12)
        ax.set_ylabel('Y坐标 (堆场长度方向)', fontsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_aspect('equal')

        power_patch = mpatches.Patch(color='lightgreen', label='有电箱区')
        no_power_patch = mpatches.Patch(color='lightcoral', label='无电箱区')
        h_track_patch = mpatches.Patch(color='red', label='横向轨道')
        v_track_patch = mpatches.Patch(color='blue', label='竖向轨道')
        ax.legend(handles=[power_patch, no_power_patch, h_track_patch, v_track_patch],
                  loc='upper left', bbox_to_anchor=(0.02, 0.98))

        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"实际堆场布局图已保存到: {save_path}")
        plt.show()

    def _draw_blocks(self, ax, blocks):
        """绘制箱区"""
        for block_id, block_info in blocks.items():
            x, y = block_info.position_x, block_info.position_y
            color = 'lightgreen' if block_info.has_power else 'lightcoral'

            rect = plt.Rectangle((x - 30, y - 30), 60, 60,
                                 facecolor=color, edgecolor='black', linewidth=1.2)
            ax.add_patch(rect)
            ax.text(x, y, block_id, ha='center', va='center', fontsize=8, fontweight='bold')

            usage_rate = block_info.current_usage / block_info.capacity
            ax.text(x, y - 15, f"{usage_rate:.1%}", ha='center', va='center',
                    fontsize=6, color='darkblue')

    def _draw_tracks(self, ax, tracks):
        """绘制轨道"""
        for track_id, track_info in tracks.items():
            start_x, start_y = track_info.start_point
            end_x, end_y = track_info.end_point

            color = 'red' if track_info.track_type == 'horizontal' else 'blue'
            ax.plot([start_x, end_x], [start_y, end_y], color=color, linewidth=2, alpha=0.8)

    def _add_area_labels(self, ax):
        """添加区域标签"""
        area_centers = {'A': (900, 500), 'B': (650, 800), 'C': (380, 1000), 'D': (180, 500)}
        for area, (x, y) in area_centers.items():
            ax.text(x, y, f"{area}区", ha='center', va='center',
                    fontsize=14, fontweight='bold', color='darkblue',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="yellow", alpha=0.7))


# ================================
# 5. RTG调度算法
# ================================
class AdvancedRTGScheduler:
    """符合实际作业逻辑的RTG调度算法"""

    def __init__(self, config: RealYardConfig):
        self.config = config
        self.tasks = []
        self.rtgs = []
        self.trucks = []
        self.blocks = {}
        self.tracks = {}
        self.track_id_to_name = {}
        self.track_name_to_id = {}
        self.BOX_LAYOUT = {}

        # RTG任务队列管理
        self.rtg_queues: Dict[int, RTGTaskQueue] = {}

    def initialize_data(self, tasks, rtgs, trucks, blocks, tracks, track_id_to_name, track_name_to_id):
        """初始化调度数据"""
        self.tasks = tasks
        self.rtgs = rtgs
        self.trucks = trucks
        self.blocks = blocks
        self.tracks = tracks
        self.track_id_to_name = track_id_to_name
        self.track_name_to_id = track_name_to_id
        self._initialize_box_layout()
        self._initialize_rtg_queues()

    def _initialize_rtg_queues(self):
        """初始化RTG任务队列"""
        for rtg in self.rtgs:
            # 获取RTG初始位置
            initial_block = self._get_rtg_initial_block(rtg)
            initial_bay = self._estimate_rtg_bay(rtg)

            self.rtg_queues[rtg.rtg_id] = RTGTaskQueue(
                rtg_id=rtg.rtg_id,
                current_block=initial_block,
                current_bay=initial_bay,
                finish_time=0.0
            )
    def _initialize_box_layout(self):
        """初始化箱区布局"""
        area_blocks = {'A': [], 'B': [], 'C': [], 'D': []}
        for block_id, block_info in self.blocks.items():
            area_blocks[block_info.area].append((block_id, block_info))

        for area, blocks_list in area_blocks.items():
            blocks_list.sort(key=lambda x: x[1].number)
            area_base_col = {'A': 0, 'B': 10, 'C': 20, 'D': 30}[area]

            for i, (block_id, block_info) in enumerate(blocks_list):
                if area == 'A':
                    row, col = i // 4, area_base_col + (i % 4)
                elif area in ['B', 'C']:
                    row, col = i // 4, area_base_col + (i % 4)
                else:  # D区
                    row, col = i // 3, area_base_col + (i % 3)

                self.BOX_LAYOUT[block_id] = {'row': row, 'col': col, 'area': area}

    def _validate_rtg_assignments(self):
        """验证RTG轨道分配（简化版）"""
        valid_tracks = set(self.tracks.keys())

        for rtg in self.rtgs:
            if rtg.current_track_id not in valid_tracks:
                # 如果当前轨道无效，从可用轨道中选择第一个
                if hasattr(rtg, 'available_tracks') and rtg.available_tracks:
                    rtg.current_track_id = rtg.available_tracks[0]
                else:
                    rtg.current_track_id = list(valid_tracks)[0] if valid_tracks else 'T001'

                print(f"为RTG-{rtg.rtg_id}重新分配轨道: {rtg.current_track_id}")
    def calculate_block_distance(self, block1, block2):
        """计算箱区距离"""
        if block1 == block2:
            return 0, 0

        if block1 not in self.BOX_LAYOUT or block2 not in self.BOX_LAYOUT:
            return 0, 0

        row1, col1 = self.BOX_LAYOUT[block1]["row"], self.BOX_LAYOUT[block1]["col"]
        row2, col2 = self.BOX_LAYOUT[block2]["row"], self.BOX_LAYOUT[block2]["col"]

        return abs(row1 - row2), abs(col1 - col2)

    def calculate_move_time(self, current_pos, target_pos, block_current, block_target):
        """计算移动时间"""
        block_row_distance, block_col_distance = self.calculate_block_distance(block_current, block_target)

        if block_row_distance == 0:
            horizontal_move_time = abs(current_pos - target_pos) *0.28
        else:
            if current_pos < 10 and target_pos < 10:
                horizontal_move_time = (current_pos + target_pos) *0.28 + 3
            elif current_pos > 10 and target_pos > 10:
                horizontal_move_time = abs(current_pos - 19) *0.28 + abs(target_pos - 19) *0.28 + 3
            elif current_pos > 10 and target_pos < 10:
                horizontal_move_time = ((19 - current_pos) + target_pos) *0.28 + 3
            else:
                horizontal_move_time = (abs(19 - target_pos) + current_pos) *0.28 + 3

        if block_col_distance == 0 or block_col_distance == 1:
            vertical_move_time = block_col_distance * 10
        else:
            vertical_move_time = 13

        return horizontal_move_time + vertical_move_time + 0.5 #加上吊具展开的时间30秒

    def _get_rtg_initial_block(self, rtg: RTGInfo) -> str:
        """获取RTG初始所在箱区"""
        # 根据轨道找到对应的箱区
        if rtg.current_track_id in self.tracks:
            track = self.tracks[rtg.current_track_id]
            if track.connected_blocks:
                return track.connected_blocks[0]
        # 如果找不到，就找第一个该RTG可达的箱区
        for block_id, block_info in self.blocks.items():
            if (block_info.area in rtg.available_areas and
                    self._can_rtg_reach_block(rtg, block_info)):
                return block_id
        # 兜底：返回任意箱区
        return list(self.blocks.keys())[0] if self.blocks else "A01"

    def _can_rtg_reach_block(self, rtg: RTGInfo, block_info: BlockInfo) -> bool:
        """判断RTG是否能到达箱区（严格按照轨道限制）"""
        # 检查区域权限
        if block_info.area not in rtg.available_areas:
            return False
        # 检查设备状态（硬性条件）
        if rtg.login_status != 'online' or rtg.fault_status != 'normal':
            return False
        # 检查轨道可达性
        if not hasattr(rtg, 'available_tracks') or not rtg.available_tracks:
            # 如果没有轨道限制，默认可达
            return True

        # 需要建立T001到实际轨道ID的映射
        rtg_actual_tracks = []
        for t_track in rtg.available_tracks:  # ['T001', 'T002'...]
            # 查找T001对应的实际轨道ID
            if t_track in self.track_name_to_id:
                actual_track_id = self.track_name_to_id[t_track]
                rtg_actual_tracks.append(actual_track_id)
            # 或者直接检查T001是否就是轨道ID
            elif t_track in self.tracks:
                rtg_actual_tracks.append(t_track)

        rtg_tracks_set = set(rtg_actual_tracks)
        block_tracks_set = set(block_info.track_connections)
        return bool(rtg_tracks_set & block_tracks_set)

    def _calculate_rtg_cost(self, rtg: RTGInfo, task: TaskInfo) -> float:
        """
        计算RTG执行任务的总成本（考虑队列等待时间）
        成本 = RTG完成当前队列任务时间 + 移动时间 + 任务执行时间
        """
        rtg_queue = self.rtg_queues[rtg.rtg_id]
        # 1. RTG完成当前所有任务的时间
        queue_finish_time = rtg_queue.get_next_available_time()
        # 2. RTG当前位置（完成最后一个任务后的位置）
        current_block, current_bay = rtg_queue.get_current_location()
        # 3. 从当前位置移动到任务位置的时间
        move_time = self.calculate_move_time(
            current_bay, task.bay, current_block, task.block_id
        )
        # 4. 任务执行时间
        task_duration = task.estimated_duration
        # 5. 总成本 = 等待时间 + 移动时间 + 执行时间
        total_cost = queue_finish_time + move_time + task_duration
        return total_cost

    def assign_rtg_to_task(self, task: TaskInfo) -> Optional[RTGInfo]:
        """为任务分配最优RTG（考虑队列等待）"""
        if task.block_id not in self.blocks:
            print(f"错误：任务{task.task_id}的箱区{task.block_id}不存在")
            return None
        block_info = self.blocks[task.block_id]
        available_rtgs = []
        # 筛选可用RTG（严格按照硬性条件）
        for rtg in self.rtgs:
            if self._can_rtg_reach_block(rtg, block_info):
                available_rtgs.append(rtg)
        if not available_rtgs:
            print(f"警告：没有可用RTG处理任务{task.task_id}（箱区{task.block_id}）")
            return None
        # 选择总成本最小的RTG
        best_rtg = min(available_rtgs, key=lambda r: self._calculate_rtg_cost(r, task))
        # 计算分配详情
        best_cost = self._calculate_rtg_cost(best_rtg, task)
        queue = self.rtg_queues[best_rtg.rtg_id]
        start_time = queue.get_next_available_time()

        print(f"任务{task.task_id} → RTG-{best_rtg.rtg_id}({best_rtg.rtg_type}), "
              f"开始时间: {start_time:.1f}分钟, 结束时间: {best_cost:.1f}分钟")
        return best_rtg

    def assign_task_to_rtg(self, task: TaskInfo, rtg: RTGInfo) -> Dict:
        """将任务分配给指定RTG并更新队列"""
        rtg_queue = self.rtg_queues[rtg.rtg_id]
        # 计算任务开始时间（等待队列完成）
        start_time = rtg_queue.get_next_available_time()
        # 计算移动时间
        current_block, current_bay = rtg_queue.get_current_location()
        move_time = self.calculate_move_time(
            current_bay, task.bay, current_block, task.block_id
        )
        # 实际开始作业时间 = 等待时间 + 移动时间
        actual_start_time = start_time + move_time
        end_time = actual_start_time + task.estimated_duration
        # 创建任务分配记录
        assignment = {
            'task_id': task.task_id,
            'rtg_id': rtg.rtg_id,
            'block_id': task.block_id,
            'bay': task.bay,
            'row': task.row,
            'tier': task.tier,
            'task_type': task.task_type,
            'estimated_time': task.estimated_duration,
            'is_cold': task.is_cold_container,
            'manual_mark': task.manual_mark,
            'queue_wait_time': start_time,
            'move_time': move_time,
            'start_time': actual_start_time,
            'end_time': end_time,
            'priority_level': self.config.TASK_PRIORITY.get(task.task_type, 99)
        }
        # 更新RTG队列
        rtg_queue.add_task(assignment)
        return assignment

    def _estimate_rtg_bay(self, rtg: RTGInfo) -> int:
        """估算RTG当前贝位"""
        return random.randint(0, 19)  # 简化处理

    def preprocess_tasks(self):
        """任务预处理"""
        manual_tasks = [task for task in self.tasks if task.manual_mark]
        normal_tasks = [task for task in self.tasks if not task.manual_mark]

        manual_tasks.sort(key=lambda t: (0 if t.is_cold_container else 1))

        def task_sort_key(task):
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            cold_priority = 0 if task.is_cold_container else 1
            return (task_priority, cold_priority)

        normal_tasks.sort(key=task_sort_key)
        return manual_tasks, normal_tasks


    def optimize_schedule(self):
        """优化调度方案（队列模式）"""
        print("\n正在优化RTG调度方案（队列模式）...")
        manual_tasks, normal_tasks = self.preprocess_tasks()

        schedule = {
            'manual_task_assignments': [],
            'normal_task_assignments': [],
            'total_makespan': 0,
            'rtg_utilization': {},
            'priority_summary': {},
            'rtg_queues': {}  # 记录每个RTG的任务队列
        }

        total_tasks = len(self.tasks)
        assigned_count = 0
        # 处理人工标记任务
        print(f"处理人工标记任务 ({len(manual_tasks)}个):")
        for i, task in enumerate(manual_tasks):
            rtg = self.assign_rtg_to_task(task)
            if rtg:
                assignment = self.assign_task_to_rtg(task, rtg)
                schedule['manual_task_assignments'].append(assignment)
                assigned_count += 1

                cold_mark = "(冷箱)" if task.is_cold_container else ""
                print(f"  {i + 1}. {task.task_id}: {task.task_type}{cold_mark}")
            else:
                print(f"  ❌ 任务{task.task_id}分配失败")
        # 处理普通任务
        print(f"处理普通任务 ({len(normal_tasks)}个):")
        current_priority = None
        task_count = 0
        for task in normal_tasks:
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            if current_priority != task_priority:
                if current_priority is not None:
                    print(f"  优先级{current_priority}完成，共{task_count}个任务")
                current_priority = task_priority
                task_count = 0
                priority_names = {1: "装卸类", 2: "收发类", 3: "翻倒类"}
                print(f"  开始处理优先级{task_priority}({priority_names.get(task_priority, '其他')}):")
            rtg = self.assign_rtg_to_task(task)
            if rtg:
                assignment = self.assign_task_to_rtg(task, rtg)
                schedule['normal_task_assignments'].append(assignment)
                assigned_count += 1
                task_count += 1
            else:
                print(f"    ❌ 任务{task.task_id}分配失败")
        if current_priority is not None:
            print(f"  优先级{current_priority}完成，共{task_count}个任务")
        # 计算最终结果
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
        # 计算makespan（所有RTG完成时间的最大值）
        if all_assignments:
            schedule['total_makespan'] = max(assignment['end_time'] for assignment in all_assignments)
        else:
            schedule['total_makespan'] = 0
        # 计算RTG利用率
        rtg_workload = {}
        for rtg_id, queue in self.rtg_queues.items():
            rtg_workload[rtg_id] = queue.finish_time
        schedule['rtg_utilization'] = rtg_workload
        # 保存RTG队列详情
        schedule['rtg_queues'] = {
            rtg_id: {
                'task_count': len(queue.tasks),
                'total_time': queue.finish_time,
                'tasks': queue.tasks
            }
            for rtg_id, queue in self.rtg_queues.items()
        }

        print(f"\n调度优化完成:")
        print(f"  - 总任务数: {total_tasks}")
        print(f"  - 成功分配: {assigned_count}个 ({assigned_count / total_tasks * 100:.1f}%)")
        print(f"  - 预计总时间: {schedule['total_makespan']:.1f}分钟")

        # # 显示RTG队列统计
        # self._print_rtg_queue_statistics()
        return schedule

    def export_schedule_to_excel(self, schedule, filename=None):
        """导出调度结果到Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"RTG调度结果_{timestamp}.xlsx"

        print(f"正在导出调度结果到: {filename}")

        wb = Workbook()
        wb.remove(wb.active)

        # 详细调度结果
        ws_detail = wb.create_sheet("详细调度结果")
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']

        headers = ['任务ID', '箱区', '贝位', '排', '层', '任务类型', '分配RTG',
                   '开始时间(分钟)', '结束时间(分钟)', '作业时长(分钟)', '是否冷箱', '优先级']
        ws_detail.append(headers)

        for assignment in all_assignments:
            task_detail = next((t for t in self.tasks if t.task_id == assignment['task_id']), None)
            row = [
                assignment['task_id'],
                assignment['block_id'],
                task_detail.bay if task_detail else '',
                task_detail.row if task_detail else '',
                task_detail.tier if task_detail else '',
                assignment['task_type'],
                f"RTG-{assignment['rtg_id']}",
                round(assignment['start_time'], 2),
                round(assignment['end_time'], 2),
                round(assignment['estimated_time'], 2),
                '是' if assignment['is_cold'] else '否',
                assignment['priority_level']
            ]
            ws_detail.append(row)

        # RTG工作负载统计
        ws_rtg = wb.create_sheet("RTG工作负载")
        ws_rtg.append(['RTG编号', 'RTG类型', '任务数量', '总工作时间', '利用率(%)'])

        rtg_stats = {}
        for assignment in all_assignments:
            rtg_id = assignment['rtg_id']
            if rtg_id not in rtg_stats:
                rtg_info = next((r for r in self.rtgs if r.rtg_id == rtg_id), None)
                rtg_stats[rtg_id] = {
                    'type': rtg_info.rtg_type if rtg_info else 'unknown',
                    'tasks': 0,
                    'time': 0
                }
            rtg_stats[rtg_id]['tasks'] += 1
            rtg_stats[rtg_id]['time'] += assignment['estimated_time']

        total_makespan = schedule['total_makespan']
        for rtg_id, stats in rtg_stats.items():
            utilization = stats['time'] / total_makespan * 100 if total_makespan > 0 else 0
            ws_rtg.append([
                f"RTG-{rtg_id}",
                stats['type'],
                stats['tasks'],
                round(stats['time'], 2),
                round(utilization, 2)
            ])

        wb.save(filename)
        print(f"调度结果已导出到: {filename}")
        return filename

# # RTG分配失败综合诊断工具
# def comprehensive_debug(scheduler):
#     """全面诊断RTG分配失败原因"""
#     print("\n" + "=" * 60)
#     print("RTG分配失败综合诊断")
#     print("=" * 60)
#
#     # 1. RTG基本状态检查
#     print("\n1. RTG状态全面检查:")
#     print("-" * 40)
#     online_normal_rtgs = []
#
#     for rtg in scheduler.rtgs:
#         status_ok = rtg.login_status == 'online' and rtg.fault_status == 'normal'
#         status_icon = "✅" if status_ok else "❌"
#
#         print(f"{status_icon} RTG-{rtg.rtg_id}({rtg.rtg_type}):")
#         print(f"    登录状态: {rtg.login_status}")
#         print(f"    故障状态: {rtg.fault_status}")
#         print(f"    区域权限: {rtg.available_areas}")
#         print(f"    轨道数量: {len(getattr(rtg, 'available_tracks', []))}")
#
#         if status_ok:
#             online_normal_rtgs.append(rtg.rtg_id)
#         print()
#
#     print(f"可用RTG总数: {len(online_normal_rtgs)}/12")
#     print(f"可用RTG列表: {online_normal_rtgs}")
#
#     # 2. 失败任务详细分析
#     print("\n2. 失败任务详细分析:")
#     print("-" * 40)
#
#     # 挑选几个典型的失败任务进行深度分析
#     failed_tasks = ['TASK004', 'TASK008', 'TASK010', 'TASK014']
#
#     for task_id in failed_tasks:
#         task = next((t for t in scheduler.tasks if t.task_id == task_id), None)
#         if not task:
#             continue
#
#         print(f"\n📋 {task_id} (箱区{task.block_id}):")
#         block_info = scheduler.blocks[task.block_id]
#
#         print(f"   箱区区域: {block_info.area}")
#         print(f"   箱区轨道: {block_info.track_connections}")
#         print(f"   轨道数量: {len(block_info.track_connections)}")
#
#         # 逐个检查每台RTG
#         can_reach_count = 0
#         for rtg in scheduler.rtgs:
#             print(f"\n   检查RTG-{rtg.rtg_id}({rtg.rtg_type}):")
#
#             # 分步检查每个条件
#             step1_area = block_info.area in rtg.available_areas
#             step2_login = rtg.login_status == 'online'
#             step3_fault = rtg.fault_status == 'normal'
#
#             print(f"     ①区域权限: {'✅' if step1_area else '❌'} "
#                   f"(需要{block_info.area}, RTG有{rtg.available_areas})")
#             print(f"     ②登录状态: {'✅' if step2_login else '❌'} ({rtg.login_status})")
#             print(f"     ③故障状态: {'✅' if step3_fault else '❌'} ({rtg.fault_status})")
#
#             # 轨道检查
#             if hasattr(rtg, 'available_tracks') and rtg.available_tracks:
#                 rtg_tracks = set(rtg.available_tracks)
#                 block_tracks = set(block_info.track_connections)
#                 intersection = rtg_tracks & block_tracks
#                 step4_track = len(intersection) > 0
#
#                 print(f"     ④轨道可达: {'✅' if step4_track else '❌'}")
#                 print(f"       RTG轨道数: {len(rtg_tracks)}")
#                 print(f"       箱区轨道数: {len(block_tracks)}")
#                 print(f"       交集: {len(intersection)}条")
#
#                 if not step4_track:
#                     # 显示RTG轨道范围和箱区轨道的具体对比
#                     block_sample = sorted(list(block_tracks))
#                     print(f"       箱区轨道: {block_sample}")
#             else:
#                 step4_track = True  # 如果没有轨道限制，默认可达
#                 print(f"     ④轨道可达: ⚠️ (无轨道限制数据)")
#
#             # 综合判断
#             all_ok = step1_area and step2_login and step3_fault and step4_track
#             if all_ok:
#                 can_reach_count += 1
#                 print(f"     ✅ 综合判断: 可用")
#             else:
#                 print(f"     ❌ 综合判断: 不可用")
#
#         print(f"\n   📊 {task_id}总结: {can_reach_count}台RTG可用")
#         if can_reach_count == 0:
#             print(f"   🚨 这就是{task_id}分配失败的原因！")
#
#     # 3. 轨道ID映射检查
#     print("\n3. 轨道ID映射检查:")
#     print("-" * 40)
#
#     # 检查RTG轨道ID与实际轨道ID是否匹配
#     all_track_ids = set(scheduler.tracks.keys())
#     print(f"系统中实际轨道ID数量: {len(all_track_ids)}")
#     print(f"轨道ID示例: {sorted(list(all_track_ids))[:10]}")
#
#     for rtg in scheduler.rtgs[:3]:  # 检查前3台RTG
#         if hasattr(rtg, 'available_tracks') and rtg.available_tracks:
#             rtg_tracks = set(rtg.available_tracks)
#             valid_tracks = rtg_tracks & all_track_ids
#             invalid_tracks = rtg_tracks - all_track_ids
#
#             print(f"\nRTG-{rtg.rtg_id}轨道匹配检查:")
#             print(f"  RTG轨道总数: {len(rtg_tracks)}")
#             print(f"  有效轨道: {len(valid_tracks)}")
#             print(f"  无效轨道: {len(invalid_tracks)}")
#
#             if invalid_tracks:
#                 print(f"  ❌ 无效轨道示例: {list(invalid_tracks)[:5]}")
#             else:
#                 print(f"  ✅ 所有轨道ID都有效")
#
#     # 4. 区域覆盖检查
#     print("\n4. 各区域RTG覆盖检查:")
#     print("-" * 40)
#
#     for area in ['A', 'B', 'C', 'D']:
#         area_blocks = [b for b_id, b in scheduler.blocks.items() if b.area == area]
#         area_rtgs = []
#
#         for rtg in scheduler.rtgs:
#             if (area in rtg.available_areas and
#                     rtg.login_status == 'online' and
#                     rtg.fault_status == 'normal'):
#                 area_rtgs.append(rtg.rtg_id)
#
#         coverage_icon = "✅" if len(area_rtgs) > 0 else "❌"
#         print(f"{coverage_icon} {area}区:")
#         print(f"    箱区数量: {len(area_blocks)}")
#         print(f"    可用RTG: {area_rtgs} ({len(area_rtgs)}台)")
#
#         if len(area_rtgs) == 0 and len(area_blocks) > 0:
#             print(f"    🚨 警告: {area}区没有任何可用RTG!")
#
# def check_specific_failure(scheduler, task_id):
#     """深度检查特定任务失败原因"""
#     task = next((t for t in scheduler.tasks if t.task_id == task_id), None)
#     if not task:
#         print(f"任务{task_id}不存在")
#         return
#
#     print(f"\n🔍 深度检查任务{task_id}:")
#     block_info = scheduler.blocks[task.block_id]
#
#     print(f"箱区信息: {task.block_id}(区域{block_info.area})")
#     print(f"函数调用测试:")
#
#     available_rtgs = []
#     for rtg in scheduler.rtgs:
#         can_reach = scheduler._can_rtg_reach_block(rtg, block_info)
#         print(f"  _can_rtg_reach_block(RTG-{rtg.rtg_id}, {task.block_id}) = {can_reach}")
#         if can_reach:
#             available_rtgs.append(rtg)
#
#     print(f"最终可用RTG: {[r.rtg_id for r in available_rtgs]}")
# ================================
# 6. 主函数
# ================================
def main():
    """主函数"""
    print("=== 基于实际堆场的RTG调度系统初始化 ===")
    config = RealYardConfig()
    data_loader = YardDataLoader(config)

    # 加载数据
    blocks, tracks, track_id_to_name, track_name_to_id = data_loader.load_yard_layout()
    tasks = data_loader.load_tasks_from_excel()
    rtgs = data_loader.load_rtgs_from_excel()
    trucks = data_loader.load_trucks(tasks)

    print(f"\n数据加载完成:")
    print(f"  - 任务数量: {len(tasks)}")
    print(f"  - RTG数量: {len(rtgs)}")
    print(f"  - 箱区数量: {len(blocks)}")
    print(f"  - 轨道数量: {len(tracks)}")

    # 统计信息
    area_stats = {}
    power_stats = {'有电': 0, '无电': 0}
    for block_info in blocks.values():
        area = block_info.area
        area_stats[area] = area_stats.get(area, 0) + 1
        if block_info.has_power:
            power_stats['有电'] += 1
        else:
            power_stats['无电'] += 1

    print(f"=== 堆场统计信息 ===")
    print("各区域箱区分布:")
    for area in sorted(area_stats.keys()):
        print(f"  {area}区: {area_stats[area]}个箱区")
    print(f"电力设施分布:")
    for power_type, count in power_stats.items():
        print(f"  {power_type}箱区: {count}个")

    # 可视化
    visualizer = RealYardVisualization(config)
    visualizer.plot_yard_layout(blocks, tracks)

    # 任务分布信息
    task_area_stats = {}
    cold_count = 0
    task_type_stats = {}

    for task in tasks:
        area = task.block_id[0] if task.block_id else 'Unknown'
        task_area_stats[area] = task_area_stats.get(area, 0) + 1

        if task.is_cold_container:
            cold_count += 1

        task_type_stats[task.task_type] = task_type_stats.get(task.task_type, 0) + 1

    print("=== 任务分布信息 ===")
    print("各区域任务分布:")
    for area in sorted(task_area_stats.keys()):
        print(f"  {area}区: {task_area_stats[area]}个任务")

    print(f"冷箱任务: {cold_count}个 ({cold_count / len(tasks) * 100:.1f}%)")
    print("任务类型分布:")
    for task_type, count in task_type_stats.items():
        print(f"  {task_type}: {count}个")

    print("=== 系统优化目标 ===")
    objectives = [
        "1. 最小化总完成时间 (makespan)",
        "2. 平衡RTG工作负载",
        "3. 减少RTG跨区域移动",
        "4. 优先处理高优先级任务",
        "5. 避免轨道拥堵",
        "6. 提高冷箱任务处理效率"
    ]
    for obj in objectives:
        print(obj)

    print("=" * 50)
    print("运行高级RTG调度算法...")

    # 运行队列调度算法
    scheduler = AdvancedRTGScheduler(config)
    scheduler.initialize_data(tasks, rtgs, trucks, blocks, tracks,
                              track_id_to_name, track_name_to_id)
    # comprehensive_debug(scheduler)
    #
    # # 🔍 深度检查特定失败任务
    # check_specific_failure(scheduler, 'TASK004')
    # check_specific_failure(scheduler, 'TASK010')
    schedule = scheduler.optimize_schedule()

    # 显示结果
    print(f"\n=== 调度结果摘要 ===")
    print(f"人工标记任务: {len(schedule['manual_task_assignments'])}个")
    print(f"普通任务分配: {len(schedule['normal_task_assignments'])}个")
    print(f"预计完成时间: {schedule['total_makespan']:.1f}分钟")

    print(f"\n=== 优先级处理统计 ===")
    priority_names = {1: "人工标记", 2: "装卸类", 3: "收发类", 4: "翻倒类"}
    for priority, stats in schedule['priority_summary'].items():
        level_name = priority_names.get(priority, f"优先级{priority}")
        print(f"  {level_name}: {stats['total']}个任务 (冷箱{stats['cold']}个)")
        for task_type, count in stats['types'].items():
            print(f"    - {task_type}: {count}个")

    print(f"\n=== RTG工作负载分布 ===")
    rtg_items = sorted(schedule['rtg_utilization'].items(), key=lambda x: x[1], reverse=True)
    for rtg_id, workload in rtg_items[:8]:
        rtg_info = next((r for r in rtgs if r.rtg_id == rtg_id), None)
        if rtg_info and workload > 0:
            utilization = workload / schedule['total_makespan'] * 100
            print(f"  RTG-{rtg_id}({rtg_info.rtg_type}): {workload:.1f}分钟 (利用率{utilization:.1f}%)")

    print(f"\n系统已完成基于实际堆场布局的RTG调度优化！")

    # 导出Excel
    try:
        excel_file = scheduler.export_schedule_to_excel(schedule)
        print(f"调度结果已导出到Excel: {excel_file}")
    except Exception as e:
        print(f"Excel导出失败: {e}")
    return tasks, rtgs, trucks, blocks, tracks, schedule


if __name__ == "__main__":
    tasks, rtgs, trucks, blocks, tracks, schedule = main()
