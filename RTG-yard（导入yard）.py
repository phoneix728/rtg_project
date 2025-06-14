import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.font_manager as fm
import time
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional, Set
import json
import warnings
import csv
from openpyxl import load_workbook

# 设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
except:
    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
warnings.filterwarnings('ignore', category=UserWarning)


# ================================
# 1. 扩展的数据结构定义
# ================================
@dataclass
class BlockInfo:
    """箱区信息数据类"""
    block_id: str  # 箱区ID，如"A01", "B05", "C12"
    area: str  # 区域，如"A", "B", "C", "D"
    number: int  # 区域内编号，如1, 5, 12
    position_x: int  # X坐标（列）
    position_y: int  # Y坐标（行）
    capacity: int  # 箱区容量
    current_usage: int  # 当前使用量
    has_power: bool  # 是否有电力设施
    track_connections: List[str]  # 连接的轨道ID列表
    bay_count: int  # 贝位数量
    row_count: int  # 排数
    tier_count: int  # 层数


@dataclass
class TrackInfo:
    """轨道信息数据类"""
    track_id: str  # 轨道ID，如"H_A_01", "V_AB_01"
    track_type: str  # 轨道类型：'horizontal'(横向) 或 'vertical'(竖向)
    start_point: Tuple[int, int]  # 起始坐标
    end_point: Tuple[int, int]  # 终止坐标
    connected_blocks: List[str]  # 连接的箱区ID列表
    track_width: float  # 轨道宽度
    max_rtg_count: int  # 最大RTG承载数量
    current_rtg_list: List[int]  # 当前在此轨道的RTG列表


@dataclass
class TaskInfo:
    """扩展的任务信息数据类"""
    task_id: str
    block_id: str  # 修改为block_id，使用新的编号系统
    bay: int  # 贝位
    row: int  # 排
    tier: int  # 层
    prep_time: str  # 准备时间
    task_type: str
    truck_id: Optional[str] = None
    is_cold_container: bool = False
    manual_mark: bool = False
    estimated_duration: float = 0.0  # 预计作业时长


@dataclass
class RTGInfo:
    """扩展的轮胎吊信息数据类"""
    rtg_id: int
    rtg_type: str  # 'electric' or 'diesel'
    current_track_id: str  # 当前所在轨道ID
    current_position: Tuple[int, int]  # 当前精确位置坐标
    login_status: str  # 状态描述
    fault_status: str  # 故障状态描述
    move_speed_horizontal: float  # 水平移动速度
    move_speed_vertical: float  # 垂直移动速度
    operation_speed: float  # 作业速度
    available_areas: List[str]  # 可作业区域列表
    max_lift_height: int  # 最大起升高度
    current_load: Optional[str] = None  # 当前承载的集装箱ID


@dataclass
class TruckInfo:
    """扩展的拖车信息数据类"""
    truck_id: str
    current_block_id: str  # 当前所在箱区
    current_bay: int  # 当前贝位
    speed: float
    containers: List[str]  # 前后箱任务ID
    arrival_time: float
    route_plan: List[str] = None  # 路线规划


# ================================
# 2. 基于实际堆场的系统配置类
# ================================
class RealYardConfig:
    """基于实际堆场布局的系统配置"""
    # GA算法参数
    POPULATION_SIZE = 150
    MAX_GENERATIONS = 300
    CROSSOVER_RATE = 0.9
    MUTATION_RATE = 0.25
    ELITE_RATE = 0.1
    # 系统规模参数
    NUM_CRANES = 12  # RTG数量
    TIME_SLOTS = [0, 120, 240, 360, 480, 600, 720]  # 时间段

    # 任务优先级（数字越小优先级越高）
    TASK_PRIORITY = {
        "装船": 1, "卸船": 1,  # 装卸优先级最高
        "进箱": 2, "提箱": 2,  # 收发箱次之
        "翻捣箱": 3  # 翻倒箱优先级最低
    }
    # 适应度权重
    WEIGHTS = {
        'makespan': 0.5, 'balance': 0.3, 'time_balance': 0.05,
        'block_switch': 0.08, 'priority_delay': 0.05,
        'conflict': 0.15, 'idle': 0.02, 'track_congestion': 0.05
    }


# ================================
# 3. 实际堆场布局生成器（整合yard.py内容）
# ================================
# 箱区坐标列表，格式: (箱区编号, 左上角X, 左上角Y, 右下角X, 右下角Y)
yard_coords = [
    # A区 (A01-09, A13)
    ("A01", 830, 100, 1000, 180),
    ("A02", 830, 200, 970, 280),
    ("A03", 830, 300, 950, 380),
    ("A04", 830, 400, 930, 480),
    ("A05", 830, 500, 910, 580),
    ("A06", 830, 600, 890, 680),
    ("A07", 830, 700, 870, 780),
    ("A08", 830, 800, 850, 880),
    ("A09", 830, 900, 850, 980),
    ("A12", 1050, 300, 1060, 1500),
    ("A13", 200, 1000, 230, 1300),
    # B区 (B01-15) - 竖向单列排列
    ("B01", 550, 100, 750, 180),
    ("B02", 550, 200, 750, 280),
    ("B03", 550, 300, 750, 380),
    ("B04", 550, 400, 750, 480),
    ("B05", 550, 500, 750, 580),
    ("B06", 550, 600, 750, 680),
    ("B07", 550, 700, 750, 780),
    ("B08", 550, 800, 750, 880),
    ("B09", 550, 900, 750, 980),
    ("B10", 550, 1000, 750, 1080),
    ("B11", 550, 1100, 750, 1180),
    ("B12", 550, 1200, 750, 1280),
    ("B13", 550, 1300, 750, 1380),
    ("B14", 550, 1400, 650, 1460),
    ("B15", 550, 1500, 600, 1620),
    # C区 (C01-20) - 竖向单列排列
    ("C01", 300, 100, 460, 180),
    ("C02", 300, 200, 460, 280),
    ("C03", 300, 300, 460, 380),
    ("C04", 300, 400, 460, 480),
    ("C05", 300, 500, 460, 580),
    ("C06", 300, 600, 460, 680),
    ("C07", 300, 700, 460, 780),
    ("C08", 300, 800, 460, 880),
    ("C09", 300, 900, 460, 980),
    ("C10", 300, 1000, 460, 1080),
    ("C11", 300, 1100, 460, 1180),
    ("C12", 300, 1200, 460, 1280),
    ("C13", 300, 1300, 460, 1380),
    ("C14", 300, 1400, 460, 1480),
    ("C16", 300, 1500, 380, 1600),
    ("C17", 300, 1620, 380, 1700),
    ("C18", 300, 1720, 440, 1800),
    ("C19", 300, 1820, 440, 1900),
    ("C20", 300, 1920, 440, 2000),
    # D区 (D14-19)
    ("D14", 140, 300, 220, 380),
    ("D15", 140, 400, 220, 480),
    ("D16", 140, 500, 220, 580),
    ("D17", 140, 600, 220, 680),
    ("D18", 140, 700, 220, 780),
    ("D19", 140, 800, 220, 880),
]
# 轨道宽度（像素单位）
TRACK_WIDTH = 5
# 轨道间距（像素单位）
TRACK_SPACING = 8


def generate_yard_model(yard_coords, track_width=TRACK_WIDTH, track_spacing=TRACK_SPACING):
    """生成堆场模型，包括箱区和轨道"""
    yards = []
    tracks = []
    track_id = 1
    for yard_id, x1, y1, x2, y2 in yard_coords:
        width = x2 - x1
        height = y2 - y1
        yards.append({
            'id': yard_id,
            'x': x1,
            'y': y1,
            'width': width,
            'height': height,
            'center_x': x1 + width / 2,
            'center_y': y1 + height / 2
        })
        # 顶部轨道：1条分为左右两段
        track_y_top = y1 - track_width - track_spacing
        # 左段
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_T_Left',
            'type': 'horizontal',
            'x1': x1,
            'y1': track_y_top,
            'x2': x1 + width / 2,
            'y2': track_y_top,
            'width': track_width
        })
        track_id += 1
        # 右段
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_T_Right',
            'type': 'horizontal',
            'x1': x1 + width / 2,
            'y1': track_y_top,
            'x2': x2,
            'y2': track_y_top,
            'width': track_width
        })
        track_id += 1
        # 底部轨道：1条分为左右两段
        track_y_bottom = y2 + track_spacing
        # 左段
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_B_Left',
            'type': 'horizontal',
            'x1': x1,
            'y1': track_y_bottom,
            'x2': x1 + width / 2,
            'y2': track_y_bottom,
            'width': track_width
        })
        track_id += 1
        # 右段
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_B_Right',
            'type': 'horizontal',
            'x1': x1 + width / 2,
            'y1': track_y_bottom,
            'x2': x2,
            'y2': track_y_bottom,
            'width': track_width
        })
        track_id += 1
        # 左侧轨道：1条完整的垂直轨道
        track_x_left = x1 - track_width
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_L',
            'type': 'vertical',
            'x1': track_x_left,
            'y1': y1,
            'x2': track_x_left,
            'y2': y2,
            'width': track_width
        })
        track_id += 1
        # 右侧轨道：1条完整的垂直轨道
        track_x_right = x2
        tracks.append({
            'id': f'T{track_id:03d}',
            'name': f'{yard_id}_R',
            'type': 'vertical',
            'x1': track_x_right,
            'y1': y1,
            'x2': track_x_right,
            'y2': y2,
            'width': track_width
        })
        track_id += 1
    return yards, tracks


def analyze_track_distribution(yards, tracks):
    """分析轨道分布统计"""
    yard_track_count = {}
    yard_track_details = {}
    for track in tracks:
        # 从轨道名称中提取箱区ID
        yard_id = track['name'].split('_')[0]
        if yard_id not in yard_track_count:
            yard_track_count[yard_id] = 0
            yard_track_details[yard_id] = []
        yard_track_count[yard_id] += 1
        yard_track_details[yard_id].append(track['name'])
    # 检查每个箱区的轨道数
    incorrect_yards = []
    for yard in yards:
        track_count = yard_track_count.get(yard['id'], 0)
        if track_count != 6:
            incorrect_yards.append(f"{yard['id']}({track_count}条)")
    # 按区域统计
    area_stats = {}
    for yard in yards:
        area = yard['id'][0]
        if area not in area_stats:
            area_stats[area] = {'yards': 0, 'tracks': 0}
        area_stats[area]['yards'] += 1
        area_stats[area]['tracks'] += yard_track_count.get(yard['id'], 0)
    return area_stats, yard_track_details


# ================================
# 4. 数据加载类 - 使用提供的Excel数据
# ================================
class YardDataLoader:
    """加载堆场数据的类"""

    def __init__(self, config: RealYardConfig):
        self.config = config

    def load_yard_layout(self):
        """加载堆场布局"""
        yards, tracks = generate_yard_model(yard_coords)
        area_stats, yard_track_details = analyze_track_distribution(yards, tracks)

        # 转换为BlockInfo和TrackInfo对象
        blocks = {}
        for yard in yards:
            block_id = yard['id']
            area = block_id[0]
            number = int(block_id[1:]) if block_id[1:].isdigit() else 0

            # 假设C区和B区前10个箱区有电力设施
            has_power = (area == 'C' and number <= 20) or (area == 'B' and number <= 10)

            blocks[block_id] = BlockInfo(
                block_id=block_id,
                area=area,
                number=number,
                position_x=int(yard['center_x']),
                position_y=int(yard['center_y']),
                capacity=180,  # 标准箱区容量
                current_usage=random.randint(40, 150),
                has_power=has_power,
                track_connections=[],
                bay_count=35,  # 35个贝位
                row_count=6,  # 6排
                tier_count=4  # 4层
            )

        # 转换轨道数据
        track_objs = {}
        for track in tracks:
            track_objs[track['id']] = TrackInfo(
                track_id=track['id'],
                track_type=track['type'],
                start_point=(track['x1'], track['y1']),
                end_point=(track['x2'], track['y2']),
                connected_blocks=[],
                track_width=track['width'],
                max_rtg_count=3 if track['type'] == 'horizontal' else 2,
                current_rtg_list=[]
            )

        # 建立箱区与轨道的连接关系
        for block_id, block_info in blocks.items():
            # 为每个箱区找到最近的轨道
            nearest_tracks = self._find_nearest_tracks(block_info, track_objs)
            for track_id in nearest_tracks:
                if track_id in track_objs:
                    block_info.track_connections.append(track_id)
                    track_objs[track_id].connected_blocks.append(block_id)

        return blocks, track_objs

    def _find_nearest_tracks(self, block_info: BlockInfo, tracks: Dict[str, TrackInfo]) -> List[str]:
        """为箱区找到最近的轨道"""
        nearest_tracks = []
        block_x, block_y = block_info.position_x, block_info.position_y

        # 检查所有轨道，找到距离在合理范围内的
        for track_id, track_info in tracks.items():
            if self._is_block_connected_to_track(block_info, track_info):
                nearest_tracks.append(track_id)
        return nearest_tracks

    def _is_block_connected_to_track(self, block_info: BlockInfo, track_info: TrackInfo) -> bool:
        """判断箱区是否连接到轨道"""
        block_x, block_y = block_info.position_x, block_info.position_y
        if track_info.track_type == 'horizontal':
            track_y = track_info.start_point[1]
            track_x_start = track_info.start_point[0]
            track_x_end = track_info.end_point[0]
            # 检查箱区是否在轨道的X范围内，且Y距离合理
            return (track_x_start <= block_x <= track_x_end and
                    abs(block_y - track_y) <= 2)
        else:
            track_x = track_info.start_point[0]
            track_y_start = track_info.start_point[1]
            track_y_end = track_info.end_point[1]
            # 检查箱区是否在轨道的Y范围内，且X距离合理
            return (track_y_start <= block_y <= track_y_end and
                    abs(block_x - track_x) <= 2)

    def load_tasks_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载任务数据"""
        wb = load_workbook(filename)
        sheet = wb['Sheet1']

        tasks = []
        # 跳过标题行
        for row in sheet.iter_rows(min_row=2):
            task_id = row[0].value
            block_id = row[1].value
            bay = int(row[2].value)
            row_num = int(row[3].value)
            tier = int(row[4].value)
            prep_time = row[5].value
            task_type = row[6].value
            truck_id = row[7].value
            is_cold = row[8].value == '是'
            manual_mark = row[9].value == '是'

            tasks.append(TaskInfo(
                task_id=task_id,
                block_id=block_id,
                bay=bay,
                row=row_num,
                tier=tier,
                prep_time=str(prep_time),
                task_type=task_type,
                truck_id=truck_id,
                is_cold_container=is_cold,
                manual_mark=manual_mark,
                estimated_duration=random.uniform(8, 25)
            ))
        return tasks

    def load_rtgs_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载RTG数据"""
        wb = load_workbook(filename)
        sheet = wb['Sheet2']

        rtgs = []
        # 假设轨道ID范围
        track_ranges = {
            'T001-T066': ['T001', 'T066'],
            'T067-T156': ['T067', 'T156'],
            'T157-T270': ['T157', 'T270'],
            'T271-T306': ['T271', 'T306']
        }

        # 跳过标题行
        for row in sheet.iter_rows(min_row=2):
            rtg_id = int(row[0].value[3:])  # 提取数字部分
            rtg_tracks = row[1].value
            rtg_type = row[2].value
            login_status = row[3].value
            fault_status = row[4].value

            # 确定当前轨道
            current_track_id = None
            if rtg_tracks in track_ranges:
                start, end = track_ranges[rtg_tracks]
                # 随机选择一个轨道
                track_nums = list(range(int(start[1:]), int(end[1:]) + 1))
                if track_nums:
                    track_num = random.choice(track_nums)
                    current_track_id = f'T{track_num:03d}'

            # 随机位置
            current_position = (0, 0)
            if current_track_id:
                # 这里需要根据轨道信息计算位置，简化处理
                current_position = (random.randint(100, 1000), random.randint(100, 2000))

            # 可用区域
            available_areas = []
            if rtg_type == 'electric':
                available_areas = ['C', 'B']
            else:
                available_areas = ['A', 'B', 'C', 'D']

            rtgs.append(RTGInfo(
                rtg_id=rtg_id,
                rtg_type=rtg_type,
                current_track_id=current_track_id,
                current_position=current_position,
                login_status=login_status,
                fault_status=fault_status,
                move_speed_horizontal=0.3,
                move_speed_vertical=0.18,
                operation_speed=2.8,
                available_areas=available_areas,
                max_lift_height=5
            ))
        return rtgs

    def load_trucks(self, tasks):
        """生成拖车数据"""
        trucks = []
        truck_tasks = {}

        for task in tasks:
            if task.truck_id:
                if task.truck_id not in truck_tasks:
                    truck_tasks[task.truck_id] = []
                truck_tasks[task.truck_id].append(task.task_id)

        for truck_id, task_ids in truck_tasks.items():
            # 随机选择一个任务的位置作为拖车当前位置
            ref_task = next(t for t in tasks if t.task_id == task_ids[0])
            truck = TruckInfo(
                truck_id=truck_id,
                current_block_id=ref_task.block_id,
                current_bay=ref_task.bay,
                speed=random.uniform(15, 22),
                containers=task_ids[:2],  # 最多2个任务
                arrival_time=random.uniform(0, 45)
            )
            trucks.append(truck)
        return trucks


# ================================
# 5. 实际堆场可视化类
# ================================
class RealYardVisualization:
    """基于实际堆场的可视化类"""

    def __init__(self, config: RealYardConfig):
        self.config = config

    def plot_yard_layout(self, blocks: Dict[str, BlockInfo], tracks: Dict[str, TrackInfo],
                         save_path: str = 'real_yard_layout.png'):
        """绘制实际堆场布局图"""
        fig, ax = plt.subplots(figsize=(18, 14))
        # 绘制箱区
        self._draw_blocks(ax, blocks)
        # 绘制轨道
        self._draw_tracks(ax, tracks)
        # 添加区域标签
        self._add_area_labels(ax)
        # 设置图例和标题
        ax.set_title('基于实际堆场的集装箱堆场布局图', fontsize=16, fontweight='bold')
        ax.set_xlabel('X坐标 (堆场宽度方向)', fontsize=12)
        ax.set_ylabel('Y坐标 (堆场长度方向)', fontsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_aspect('equal')
        # 添加图例
        power_patch = mpatches.Patch(color='lightgreen', label='有电箱区')
        no_power_patch = mpatches.Patch(color='lightcoral', label='无电箱区')
        h_track_patch = mpatches.Patch(color='red', label='横向轨道')
        v_track_patch = mpatches.Patch(color='blue', label='竖向轨道')
        main_track_patch = mpatches.Patch(color='darkred', label='主要轨道')
        ax.legend(handles=[power_patch, no_power_patch, h_track_patch, v_track_patch, main_track_patch],
                  loc='upper left', bbox_to_anchor=(0.02, 0.98))
        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"实际堆场布局图已保存到: {save_path}")
        plt.show()

    def _draw_blocks(self, ax, blocks: Dict[str, BlockInfo]):
        """绘制箱区"""
        for block_id, block_info in blocks.items():
            x = block_info.position_x
            y = block_info.position_y
            # 根据是否有电选择颜色
            color = 'lightgreen' if block_info.has_power else 'lightcoral'
            # 绘制箱区矩形（缩小显示）
            rect = plt.Rectangle((x - 20, y - 20), 40, 40,
                                 facecolor=color, edgecolor='black', linewidth=1.2)
            ax.add_patch(rect)
            # 添加箱区标签
            ax.text(x, y, block_id, ha='center', va='center',
                    fontsize=7, fontweight='bold')
            # 添加使用率信息
            usage_rate = block_info.current_usage / block_info.capacity
            usage_text = f"{usage_rate:.1%}"
            ax.text(x, y - 10, usage_text, ha='center', va='center',
                    fontsize=5, color='darkblue')

    def _draw_tracks(self, ax, tracks: Dict[str, TrackInfo]):
        """绘制轨道"""
        for track_id, track_info in tracks.items():
            start_x, start_y = track_info.start_point
            end_x, end_y = track_info.end_point
            # 根据轨道类型和重要性选择颜色和宽度
            if 'MAIN' in track_id or 'SEASIDE' in track_id or 'LANDSIDE' in track_id:
                color = 'darkred'
                linewidth = 4
            elif track_info.track_type == 'horizontal':
                color = 'red'
                linewidth = 2
            else:
                color = 'blue'
                linewidth = 2
            # 绘制轨道线
            ax.plot([start_x, end_x], [start_y, end_y],
                    color=color, linewidth=linewidth, alpha=0.8)

    def _add_area_labels(self, ax):
        """添加区域标签"""
        area_centers = {
            'A': (900, 500),
            'B': (650, 800),
            'C': (380, 1000),
            'D': (180, 500)
        }
        for area, (x, y) in area_centers.items():
            ax.text(x, y, f"{area}区", ha='center', va='center',
                    fontsize=14, fontweight='bold', color='darkblue',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="yellow", alpha=0.7))


# ================================
# 6. 使用示例
# ================================
def main():
    """主函数示例"""
    print("=== 基于实际堆场的RTG调度系统初始化 ===")
    # 初始化配置
    config = RealYardConfig()
    # 初始化数据加载器
    data_loader = YardDataLoader(config)

    # 加载堆场布局
    blocks, tracks = data_loader.load_yard_layout()
    # 加载任务数据
    tasks = data_loader.load_tasks_from_excel()
    # 加载RTG数据
    rtgs = data_loader.load_rtgs_from_excel()
    # 加载拖车数据
    trucks = data_loader.load_trucks(tasks)

    print(f"基于实际堆场的数据加载完成:")
    print(f"  - 任务数量: {len(tasks)}")
    print(f"  - RTG数量: {len(rtgs)}")
    print(f"  - 拖车数量: {len(trucks)}")
    print(f"  - 箱区数量: {len(blocks)}")
    print(f"  - 轨道数量: {len(tracks)}")

    # 统计各区域箱区数量
    area_stats = {}
    power_stats = {'有电': 0, '无电': 0}
    for block_id, block_info in blocks.items():
        area = block_info.area
        if area not in area_stats:
            area_stats[area] = 0
        area_stats[area] += 1
        if block_info.has_power:
            power_stats['有电'] += 1
        else:
            power_stats['无电'] += 1

    print(f"\n=== 堆场统计信息 ===")
    print("各区域箱区分布:")
    for area, count in sorted(area_stats.items()):
        print(f"  {area}区: {count}个箱区")
    print(f"电力设施分布:")
    print(f"  有电箱区: {power_stats['有电']}个")
    print(f"  无电箱区: {power_stats['无电']}个")

    # 可视化布局
    visualizer = RealYardVisualization(config)
    visualizer.plot_yard_layout(blocks, tracks)

    # 显示部分详细数据
    print("\n=== 箱区信息样例 ===")
    for i, (block_id, block_info) in enumerate(list(blocks.items())[:8]):
        power_status = "有电" if block_info.has_power else "无电"
        usage_rate = block_info.current_usage / block_info.capacity
        track_count = len(block_info.track_connections)
        print(f"{block_info.block_id}: {block_info.area}区, 位置({block_info.position_x},{block_info.position_y}), "
              f"{power_status}, 使用率{usage_rate:.1%}, 连接{track_count}条轨道")

    print("\n=== 轨道信息样例 ===")
    for i, (track_id, track_info) in enumerate(list(tracks.items())[:6]):
        track_type_cn = "横向" if track_info.track_type == "horizontal" else "竖向"
        connected_count = len(track_info.connected_blocks)
        rtg_count = len(track_info.current_rtg_list)
        print(f"{track_info.track_id}: {track_type_cn}轨道, "
              f"连接{connected_count}个箱区, 当前{rtg_count}台RTG")

    print("\n=== RTG配置信息 ===")
    electric_count = sum(1 for rtg in rtgs if rtg.rtg_type == 'electric')
    diesel_count = sum(1 for rtg in rtgs if rtg.rtg_type == 'diesel')
    print(f"电动RTG: {electric_count}台 (主要服务有电箱区)")
    print(f"柴油RTG: {diesel_count}台 (服务所有箱区)")

    # RTG分布统计
    rtg_by_track = {}
    for rtg in rtgs:
        track = rtg.current_track_id
        if track and track in tracks:
            if track not in rtg_by_track:
                rtg_by_track[track] = []
            rtg_by_track[track].append(rtg.rtg_id)

    print(f"\nRTG轨道分布:")
    for track_id, rtg_list in list(rtg_by_track.items())[:5]:
        print(f"  {track_id}: RTG {rtg_list}")

    # 任务分布统计
    task_by_area = {}
    cold_container_count = 0
    for task in tasks:
        block_info = blocks[task.block_id]
        area = block_info.area
        if area not in task_by_area:
            task_by_area[area] = 0
        task_by_area[area] += 1
        if task.is_cold_container:
            cold_container_count += 1

    print(f"\n=== 任务分布信息 ===")
    print("各区域任务分布:")
    for area, count in sorted(task_by_area.items()):
        print(f"  {area}区: {count}个任务")
    print(f"冷箱任务: {cold_container_count}个 ({cold_container_count / len(tasks):.1%})")

    # 任务类型统计
    task_type_stats = {}
    for task in tasks:
        task_type = task.task_type
        if task_type not in task_type_stats:
            task_type_stats[task_type] = 0
        task_type_stats[task_type] += 1

    print(f"\n任务类型分布:")
    for task_type, count in task_type_stats.items():
        print(f"  {task_type}: {count}个")

    print(f"\n=== 系统优化目标 ===")
    print("1. 最小化总完成时间 (makespan)")
    print("2. 平衡RTG工作负载")
    print("3. 减少RTG跨区域移动")
    print("4. 优先处理高优先级任务")
    print("5. 避免轨道拥堵")
    print("6. 提高冷箱任务处理效率")

    return tasks, rtgs, trucks, blocks, tracks


# ================================
# 7. 高级调度算法框架
# ================================
class AdvancedRTGScheduler:
    """高级RTG调度算法"""

    def __init__(self, config: RealYardConfig):
        self.config = config
        self.tasks = []
        self.rtgs = []
        self.trucks = []
        self.blocks = {}
        self.tracks = {}

    def initialize_data(self, tasks, rtgs, trucks, blocks, tracks):
        """初始化调度数据"""
        self.tasks = tasks
        self.rtgs = rtgs
        self.trucks = trucks
        self.blocks = blocks
        self.tracks = tracks

    def preprocess_tasks(self):
        """任务预处理 - 按正确优先级排序"""
        print("正在进行任务预处理...")
        # 1. 首先按是否人工标记分类
        manual_tasks = [task for task in self.tasks if task.manual_mark]
        normal_tasks = [task for task in self.tasks if not task.manual_mark]

        # 2. 对人工标记任务排序（冷箱优先）
        manual_tasks.sort(key=lambda t: (0 if t.is_cold_container else 1))

        # 3. 对普通任务排序：先按任务类型优先级，再按是否冷箱
        def task_sort_key(task):
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            cold_priority = 0 if task.is_cold_container else 1
            return (task_priority, cold_priority)

        normal_tasks.sort(key=task_sort_key)

        # 4. 按任务类型分组统计
        task_type_stats = {}
        cold_count = {'manual': 0, 'normal': 0}
        for task in self.tasks:
            task_type = task.task_type
            if task_type not in task_type_stats:
                task_type_stats[task_type] = {'manual': 0, 'normal': 0, 'cold': 0}
            if task.manual_mark:
                task_type_stats[task_type]['manual'] += 1
                if task.is_cold_container:
                    cold_count['manual'] += 1
            else:
                task_type_stats[task_type]['normal'] += 1
                if task.is_cold_container:
                    cold_count['normal'] += 1
            if task.is_cold_container:
                task_type_stats[task_type]['cold'] += 1

        print(f"  - 人工标记任务: {len(manual_tasks)}个 (其中冷箱 {cold_count['manual']}个)")
        print(f"  - 普通任务: {len(normal_tasks)}个 (其中冷箱 {cold_count['normal']}个)")
        print(f"  任务类型优先级分布:")
        for task_type in ["装船", "卸船", "进箱", "提箱", "翻捣箱"]:
            if task_type in task_type_stats:
                stats = task_type_stats[task_type]
                priority = self.config.TASK_PRIORITY.get(task_type, 99)
                print(
                    f"    {task_type}(优先级{priority}): 普通{stats['normal']}个, 人工{stats['manual']}个, 冷箱{stats['cold']}个")

        return manual_tasks, normal_tasks

    def assign_rtg_to_task(self, task: TaskInfo) -> Optional[RTGInfo]:
        """为任务分配最优RTG"""
        block_info = self.blocks[task.block_id]
        available_rtgs = []

        # 筛选可用的RTG
        for rtg in self.rtgs:
            if (rtg.login_status == 'online' and
                    rtg.fault_status == 'normal' and
                    block_info.area in rtg.available_areas):
                # 冷箱任务优先使用电动RTG（如果有电设施）
                if task.is_cold_container and block_info.has_power:
                    if rtg.rtg_type == 'electric':
                        available_rtgs.insert(0, rtg)  # 电动RTG排在前面
                    else:
                        available_rtgs.append(rtg)  # 柴油RTG作为备选
                else:
                    available_rtgs.append(rtg)

        if not available_rtgs:
            return None

        # 选择距离最近的RTG
        best_rtg = min(available_rtgs,
                       key=lambda r: self._calculate_distance(r, block_info))
        return best_rtg

    def _calculate_distance(self, rtg: RTGInfo, block_info: BlockInfo) -> float:
        """计算RTG到箱区的距离"""
        rtg_x, rtg_y = rtg.current_position
        block_x, block_y = block_info.position_x, block_info.position_y
        return ((rtg_x - block_x) ** 2 + (rtg_y - block_y) ** 2) ** 0.5

    def optimize_schedule(self):
        """优化调度方案 - 按正确优先级处理"""
        print("\n正在优化RTG调度方案...")
        print("优先级规则: 人工标记 > 普通任务(装卸>收发>翻倒箱) > 同优先级内冷箱>普通箱")
        manual_tasks, normal_tasks = self.preprocess_tasks()

        schedule = {
            'manual_task_assignments': [],
            'normal_task_assignments': [],
            'total_makespan': 0,
            'rtg_utilization': {},
            'priority_summary': {}
        }

        # 记录当前RTG工作负载
        rtg_workload = {rtg.rtg_id: 0 for rtg in self.rtgs}
        rtg_schedule = {rtg.rtg_id: [] for rtg in self.rtgs}

        # 1. 优先处理人工标记任务（最高优先级）
        print(f"\n处理人工标记任务 ({len(manual_tasks)}个):")
        for i, task in enumerate(manual_tasks):
            rtg = self.assign_rtg_to_task(task)
            if rtg:
                start_time = rtg_workload[rtg.rtg_id]
                assignment = {
                    'task_id': task.task_id,
                    'rtg_id': rtg.rtg_id,
                    'block_id': task.block_id,
                    'task_type': task.task_type,
                    'estimated_time': task.estimated_duration,
                    'is_cold': task.is_cold_container,
                    'priority_level': 1,
                    'start_time': start_time,
                    'end_time': start_time + task.estimated_duration
                }
                schedule['manual_task_assignments'].append(assignment)
                rtg_workload[rtg.rtg_id] += task.estimated_duration
                rtg_schedule[rtg.rtg_id].append(assignment)
                cold_mark = "(冷箱)" if task.is_cold_container else ""
                print(
                    f"  {i + 1}. {task.task_id}: {task.task_type}{cold_mark} -> RTG-{rtg.rtg_id} ({start_time:.1f}-{start_time + task.estimated_duration:.1f}分钟)")

        # 2. 处理普通任务（按装卸>收发>翻倒箱优先级，同级内冷箱优先）
        print(f"\n处理普通任务 ({len(normal_tasks)}个):")
        # 按优先级分组显示
        current_priority = None
        task_count = 0
        for i, task in enumerate(normal_tasks):
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            if current_priority != task_priority:
                if current_priority is not None:
                    print(f"  优先级{current_priority}完成，共{task_count}个任务")
                current_priority = task_priority
                task_count = 0
                priority_name = {1: "装卸类", 2: "收发类", 3: "翻倒类"}.get(task_priority, "其他")
                print(f"  开始处理优先级{task_priority}({priority_name}):")

            rtg = self.assign_rtg_to_task(task)
            if rtg:
                start_time = rtg_workload[rtg.rtg_id]
                assignment = {
                    'task_id': task.task_id,
                    'rtg_id': rtg.rtg_id,
                    'block_id': task.block_id,
                    'task_type': task.task_type,
                    'estimated_time': task.estimated_duration,
                    'is_cold': task.is_cold_container,
                    'priority_level': task_priority + 1,  # +1因为人工标记是1
                    'start_time': start_time,
                    'end_time': start_time + task.estimated_duration
                }
                schedule['normal_task_assignments'].append(assignment)
                rtg_workload[rtg.rtg_id] += task.estimated_duration
                rtg_schedule[rtg.rtg_id].append(assignment)
                cold_mark = "(冷箱)" if task.is_cold_container else ""
                print(f"    {task.task_id}: {task.task_type}{cold_mark} -> RTG-{rtg.rtg_id}")
                task_count += 1

        if current_priority is not None:
            print(f"  优先级{current_priority}完成，共{task_count}个任务")

        # 计算调度结果
        schedule['rtg_utilization'] = rtg_workload
        schedule['total_makespan'] = max(rtg_workload.values()) if rtg_workload.values() else 0

        # 统计优先级分布
        priority_stats = {}
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
        for assignment in all_assignments:
            priority = assignment['priority_level']
            task_type = assignment['task_type']
            is_cold = assignment['is_cold']
            if priority not in priority_stats:
                priority_stats[priority] = {'total': 0, 'cold': 0, 'types': {}}
            priority_stats[priority]['total'] += 1
            if is_cold:
                priority_stats[priority]['cold'] += 1
            if task_type not in priority_stats[priority]['types']:
                priority_stats[priority]['types'][task_type] = 0
            priority_stats[priority]['types'][task_type] += 1

        schedule['priority_summary'] = priority_stats

        print(f"\n调度优化完成:")
        print(f"  - 总任务数: {len(all_assignments)}")
        print(f"  - 预计总时间: {schedule['total_makespan']:.1f}分钟")
        avg_utilization = sum(rtg_workload.values()) / len(rtg_workload) / schedule['total_makespan'] * 100 if schedule[
                                                                                                                   'total_makespan'] > 0 else 0
        print(f"  - 平均RTG利用率: {avg_utilization:.1f}%")

        return schedule


if __name__ == "__main__":
    # 运行主程序
    tasks, rtgs, trucks, blocks, tracks = main()

    # 运行高级调度算法示例
    print("\n" + "=" * 50)
    print("运行高级RTG调度算法...")
    scheduler = AdvancedRTGScheduler(RealYardConfig())
    scheduler.initialize_data(tasks, rtgs, trucks, blocks, tracks)
    schedule = scheduler.optimize_schedule()

    print(f"\n=== 调度结果摘要 ===")
    print(f"人工标记任务: {len(schedule['manual_task_assignments'])}个")
    print(f"普通任务分配: {len(schedule['normal_task_assignments'])}个")
    print(f"预计完成时间: {schedule['total_makespan']:.1f}分钟")

    # 显示优先级统计
    print(f"\n=== 优先级处理统计 ===")
    for priority, stats in schedule['priority_summary'].items():
        if priority == 1:
            level_name = "人工标记"
        elif priority == 2:
            level_name = "装卸类"
        elif priority == 3:
            level_name = "收发类"
        elif priority == 4:
            level_name = "翻倒类"
        else:
            level_name = f"优先级{priority}"
        print(f"  {level_name}: {stats['total']}个任务 (冷箱{stats['cold']}个)")
        for task_type, count in stats['types'].items():
            print(f"    - {task_type}: {count}个")

    # 显示RTG工作负载分布
    print(f"\n=== RTG工作负载分布 ===")
    rtg_items = sorted(schedule['rtg_utilization'].items(), key=lambda x: x[1], reverse=True)
    for rtg_id, workload in rtg_items[:8]:  # 显示前8台最忙的RTG
        rtg_info = next(r for r in rtgs if r.rtg_id == rtg_id)
        utilization = workload / schedule['total_makespan'] * 100 if schedule['total_makespan'] > 0 else 0
        print(f"  RTG-{rtg_id}({rtg_info.rtg_type}): {workload:.1f}分钟 (利用率{utilization:.1f}%)")

    # 显示冷箱处理统计
    total_cold = sum(1 for t in tasks if t.is_cold_container)
    manual_cold = len([a for a in schedule['manual_task_assignments'] if a['is_cold']])
    normal_cold = len([a for a in schedule['normal_task_assignments'] if a['is_cold']])

    print(f"\n=== 冷箱处理统计 ===")
    print(f"总冷箱任务: {total_cold}个")
    print(f"  - 人工标记冷箱: {manual_cold}个")
    print(f"  - 普通冷箱: {normal_cold}个")

    # 按任务类型统计冷箱
    cold_by_type = {}
    all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
    for assignment in all_assignments:
        if assignment['is_cold']:
            task_type = assignment['task_type']
            if task_type not in cold_by_type:
                cold_by_type[task_type] = 0
            cold_by_type[task_type] += 1

    print(f"  冷箱任务类型分布:")
    for task_type in ["装船", "卸船", "进箱", "提箱", "翻捣箱"]:
        if task_type in cold_by_type:
            print(f"    - {task_type}: {cold_by_type[task_type]}个")

    print(f"\n系统已完成基于实际堆场布局的RTG调度优化！")