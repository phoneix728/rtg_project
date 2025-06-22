import numpy as np
import random
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import warnings
from openpyxl import Workbook
from datetime import datetime
import re
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Set
import heapq
from openpyxl import load_workbook

# 导入yard.py的功能
from yard import generate_yard_model, yard_coords, analyze_track_distribution, TRACK_WIDTH, TRACK_SPACING

# 设置中文字体
try:
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
except:
    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans']
warnings.filterwarnings('ignore', category=UserWarning)


# ================================
# 1. 数据结构定义
# ================================
@dataclass
class BlockInfo:
    """箱区信息数据类"""
    block_id: str
    area: str
    number: int
    position_x: int
    position_y: int
    capacity: int
    current_usage: int
    has_power: bool
    track_connections: List[str]
    bay_count: int
    row_count: int
    tier_count: int


@dataclass
class TrackInfo:
    """轨道信息数据类"""
    track_id: str
    track_type: str
    start_point: Tuple[int, int]
    end_point: Tuple[int, int]
    connected_blocks: List[str]
    track_width: float
    max_rtg_count: int
    current_rtg_list: List[int]


@dataclass
class TaskInfo:
    """任务信息数据类"""
    task_id: str
    block_id: str
    bay: int
    row: int
    tier: int
    truck_arrival_time: float  # 拖车到达时间（分钟）
    task_type: str
    truck_id: Optional[str] = None
    deadline: Optional[float] = None  # 时间承诺截止时间（分钟）
    container_size: int = 40  # 集装箱尺寸
    is_cold_container: bool = False
    manual_mark: bool = False
    estimated_duration: float = 0.0


@dataclass
class RTGInfo:
    """轮胎吊信息数据类"""
    rtg_id: str  # RTG编号（字符串）
    rtg_tracks: str  # RTG轨道范围
    rtg_type: str  # RTG类型：diesel/electric
    login_status: str
    fault_status: str
    current_bay: int
    has_container: bool = False
    move_speed_horizontal: float = 0.3
    move_speed_vertical: float = 0.18
    operation_speed: float = 2.8
    available_areas: List[str] = field(default_factory=list)
    available_tracks: List[str] = field(default_factory=list)
    max_lift_height: int = 6


@dataclass
class TruckInfo:
    """拖车信息数据类"""
    truck_id: str
    current_block_id: str
    current_bay: int
    speed: float
    containers: List[str]
    arrival_time: float
    route_plan: List[str] = None


@dataclass
class RTGTaskQueue:
    """RTG任务队列"""
    rtg_id: str  # 改为字符串
    tasks: List[Dict] = field(default_factory=list)  # 任务队列
    current_position: Tuple[int, int] = (0, 0)  # 当前位置
    current_block: str = ""  # 当前箱区
    current_bay: int = 0  # 当前贝位
    finish_time: float = 0.0  # 当前所有任务完成时间

    def add_task(self, task_assignment: Dict):
        """添加任务到队列"""
        self.tasks.append(task_assignment)
        self.finish_time = task_assignment['end_time']
        self.current_block = task_assignment['block_id']
        self.current_bay = task_assignment['bay']

    def get_next_available_time(self) -> float:
        """获取RTG下次可用时间"""
        return self.finish_time

    def get_current_location(self) -> Tuple[str, int]:
        """获取RTG当前位置（箱区，贝位）"""
        if not self.tasks:
            # 如果没有任务，返回RTG初始位置
            return self.current_block, self.current_bay
        else:
            # 返回最后一个任务的位置
            last_task = self.tasks[-1]
            return last_task['block_id'], last_task['bay']


# ================================
# 2. 系统配置类
# ================================
class RealYardConfig:
    """系统配置"""
    POPULATION_SIZE = 150
    MAX_GENERATIONS = 300
    CROSSOVER_RATE = 0.9
    MUTATION_RATE = 0.25
    ELITE_RATE = 0.1
    NUM_CRANES = 12
    TIME_SLOTS = [0, 120, 240, 360, 480, 600, 720]

    TASK_PRIORITY = {
        "装船": 1, "卸船": 1,
        "进箱": 2, "提箱": 2,
        "翻捣箱": 3
    }

    WEIGHTS = {
        'makespan': 0.5, 'balance': 0.3, 'time_balance': 0.05,
        'block_switch': 0.08, 'priority_delay': 0.05,
        'conflict': 0.15, 'idle': 0.02, 'track_congestion': 0.05
    }


# ================================
# 3. 数据加载类
# ================================
class YardDataLoader:
    """数据加载器 """

    def __init__(self, config: RealYardConfig):
        self.config = config

    def parse_datetime(self, datetime_str: str) -> float:
        """解析时间字符串为分钟数"""
        try:
            if isinstance(datetime_str, str) and datetime_str.strip():
                # 尝试多种时间格式
                formats = [
                    "%Y/%m/%d %H:%M",
                    "%Y-%m-%d %H:%M",
                    "%Y/%m/%d %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S"
                ]

                for fmt in formats:
                    try:
                        dt = datetime.strptime(datetime_str.strip(), fmt)
                        base_time = datetime.strptime(dt.strftime("%Y/%m/%d 00:00"), "%Y/%m/%d %H:%M")
                        minutes = (dt - base_time).total_seconds() / 60
                        print(f"成功解析时间: {datetime_str} -> {minutes:.1f}分钟")
                        return minutes
                    except ValueError:
                        continue

                print(f"警告：无法解析时间格式: {datetime_str}")
                return 0.0
            elif datetime_str is None:
                return None
            else:
                return 0.0
        except Exception as e:
            print(f"时间解析错误: {datetime_str}, 错误: {e}")
            return 0.0

    def load_yard_layout(self):
        """加载堆场布局"""
        print("正在加载堆场布局...")
        yards, tracks = generate_yard_model(yard_coords, TRACK_WIDTH, TRACK_SPACING)
        # 处理analyze_track_distribution的返回值
        try:
            analysis_result = analyze_track_distribution(yards, tracks)
            if analysis_result is not None:
                area_stats, yard_track_details = analysis_result
            else:
                area_stats, yard_track_details = {}, {}
        except Exception as e:
            print(f"轨道分析失败: {e}")
            area_stats, yard_track_details = {}, {}

        # 转换为BlockInfo对象
        blocks = {}
        for yard in yards:
            block_id = yard.get('id', 'UNKNOWN')
            area = block_id[0] if block_id != 'UNKNOWN' else 'A'
            number_str = block_id[1:] if len(block_id) > 1 else '0'
            number = int(number_str) if number_str.isdigit() else 0
            has_power = (area == 'C' and number <= 20) or (area == 'B' and number <= 10)

            blocks[block_id] = BlockInfo(
                block_id=block_id,
                area=area,
                number=number,
                position_x=int(yard.get('center_x', 0)),
                position_y=int(yard.get('center_y', 0)),
                capacity=180,
                current_usage=random.randint(40, 150),
                has_power=has_power,
                track_connections=[],  # 稍后建立连接关系
                bay_count=19,
                row_count=6,
                tier_count=6
            )

        # 转换轨道数据并建立ID映射
        track_objs = {}
        track_id_to_name = {}  # track_id到track_name的映射
        track_name_to_id = {}  # track_name到track_id的映射

        for i, track in enumerate(tracks):
            track_id = track.get('id', f'T{i + 1:03d}')
            track_name = track.get('name', track_id)  # 从yard.py获取的轨道名称

            track_objs[track_id] = TrackInfo(
                track_id=track_id,
                track_type=track.get('type', 'horizontal'),
                start_point=(track.get('x1', 0), track.get('y1', 0)),
                end_point=(track.get('x2', 0), track.get('y2', 0)),
                connected_blocks=[],
                track_width=track.get('width', TRACK_WIDTH),
                max_rtg_count=3 if track.get('type', 'horizontal') == 'horizontal' else 2,
                current_rtg_list=[]
            )

            # 建立ID和名称的双向映射
            track_id_to_name[track_id] = track_name
            track_name_to_id[track_name] = track_id

        # 建立箱区与轨道连接关系（新逻辑）
        self._establish_block_track_connections(blocks, track_objs, track_name_to_id)

        print(f"堆场布局加载完成：{len(blocks)}个箱区，{len(track_objs)}条轨道")
        return blocks, track_objs, track_id_to_name, track_name_to_id

    def _establish_block_track_connections(self, blocks, tracks, track_name_to_id):
        """
        建立箱区与轨道的连接关系
        直接从yard.py读取的轨道name中解析箱区ID
        """
        # 初始化所有箱区的轨道连接列表
        for block_id in blocks.keys():
            blocks[block_id].track_connections = []
        # 遍历所有轨道，根据name建立连接
        for track_name, track_id in track_name_to_id.items():
            # 从轨道名称中提取箱区ID（例如：A01_L1 -> A01）
            if '_' in track_name:
                block_id = track_name.split('_')[0]  # 取下划线前的部分作为箱区ID
                # 如果该箱区存在，建立连接关系
                if block_id in blocks:
                    blocks[block_id].track_connections.append(track_id)
                    tracks[track_id].connected_blocks.append(block_id)
                else:
                    print(f"警告：轨道{track_name}对应的箱区{block_id}不存在")

    def _parse_rtg_track_range(self, rtg_tracks_str):
        """
        解析RTG轨道范围字符串
        例如：'T001-T066' -> ['T001', 'T002', ..., 'T066']
        """
        if not rtg_tracks_str or rtg_tracks_str == 'None':
            return []
        try:
            # 匹配 T001-T066 格式
            match = re.match(r'T(\d+)-T(\d+)', rtg_tracks_str)
            if match:
                start_num = int(match.group(1))
                end_num = int(match.group(2))
                return [f'T{i:03d}' for i in range(start_num, end_num + 1)]
            else:
                # 如果不是范围格式，尝试作为单个轨道ID处理
                return [rtg_tracks_str] if rtg_tracks_str.startswith('T') else []
        except Exception as e:
            print(f"解析RTG轨道范围失败: {rtg_tracks_str}, 错误: {e}")
            return []

    def load_tasks_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载任务数据（新格式）"""
        wb = load_workbook(filename)
        sheet = wb['Sheet1']
        tasks = []

        # 获取表头
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 如果第一列为空，跳过
                continue

            # 创建字典映射
            row_data = dict(zip(headers, row))

            # 解析拖车到达时间
            truck_arrival_raw = row_data.get('truck_arrival_time', '')
            truck_arrival_time = self.parse_datetime(str(truck_arrival_raw)) if truck_arrival_raw else 0.0

            # 解析deadline（时间承诺）
            deadline_raw = row_data.get('deadline', '')
            deadline = None
            if deadline_raw and str(deadline_raw).strip() and str(deadline_raw).strip().lower() != 'none':
                deadline = self.parse_datetime(str(deadline_raw))

            tasks.append(TaskInfo(
                task_id=str(row_data['task_id']),
                block_id=str(row_data['block_id']),
                bay=int(row_data.get('bay', 0)),
                row=int(row_data.get('row', 0)),
                tier=int(row_data.get('tier', 0)),
                truck_arrival_time=truck_arrival_time,  # 使用拖车到达时间
                task_type=str(row_data.get('task_type', '')),
                truck_id=str(row_data.get('truck_id', '')) if row_data.get('truck_id') else None,
                deadline=deadline,  # 直接读取deadline
                container_size=int(row_data.get('container_size', 40)),
                is_cold_container=str(row_data.get('is_cold_container', '否')) == '是',
                manual_mark=str(row_data.get('manual_mark', '否')) == '是',
                estimated_duration=random.uniform(8, 25)
            ))
        print(f"从Excel加载了{len(tasks)}个任务")
        return tasks

    def load_rtgs_from_excel(self, filename='code数据.xlsx'):
        """从Excel加载RTG数据（新格式）"""
        wb = load_workbook(filename)
        sheet = wb['Sheet2']
        rtgs = []

        # 获取表头
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 如果第一列为空，跳过
                continue

            # 创建字典映射
            row_data = dict(zip(headers, row))

            rtg_id = str(row_data['rtg_id'])
            rtg_tracks = str(row_data.get('rtg_tracks', ''))
            rtg_type = str(row_data.get('rtg_type', 'diesel')).lower()

            # 解析RTG可操作的轨道范围
            rtg_available_tracks = self._parse_rtg_track_range(rtg_tracks)

            # 根据RTG类型确定可用区域
            if rtg_type == 'electric':
                available_areas = ['C', 'B']  # 电动RTG只能在有电区域
            else:
                available_areas = ['A', 'B', 'C', 'D']  # 柴油RTG可以在所有区域

            rtg_info = RTGInfo(
                rtg_id=rtg_id,
                rtg_tracks=rtg_tracks,
                rtg_type=rtg_type,
                login_status=str(row_data.get('login_status', 'offline')),
                fault_status=str(row_data.get('fault_status', 'normal')),
                current_bay=int(row_data.get('current_bay', 0)),
                has_container=str(row_data.get('has_container', '否')) == '是',
                available_areas=available_areas,
                available_tracks=rtg_available_tracks
            )

            rtgs.append(rtg_info)

        print(f"从Excel加载了{len(rtgs)}台RTG")
        return rtgs

    def load_trucks(self, tasks):
        """生成拖车数据"""
        trucks = []
        truck_tasks = {}

        for task in tasks:
            if task.truck_id:
                if task.truck_id not in truck_tasks:
                    truck_tasks[task.truck_id] = []
                truck_tasks[task.truck_id].append(task.task_id)

        for truck_id, task_ids in truck_tasks.items():
            ref_task = next(t for t in tasks if t.task_id == task_ids[0])
            trucks.append(TruckInfo(
                truck_id=truck_id,
                current_block_id=ref_task.block_id,
                current_bay=ref_task.bay,
                speed=random.uniform(15, 22),
                containers=task_ids[:2],
                arrival_time=random.uniform(0, 45)
            ))
        print(f"生成了{len(trucks)}台拖车")
        return trucks


# ================================
# 4. 可视化类
# ================================
class RealYardVisualization:
    """可视化类"""

    def __init__(self, config: RealYardConfig):
        self.config = config

    def plot_yard_layout(self, blocks: Dict[str, BlockInfo], tracks: Dict[str, TrackInfo],
                         save_path: str = 'real_yard_layout.png'):
        """绘制堆场布局图"""
        fig, ax = plt.subplots(figsize=(18, 14))

        self._draw_blocks(ax, blocks)
        self._draw_tracks(ax, tracks)
        self._add_area_labels(ax)

        ax.set_title('基于yard.py的集装箱堆场布局图', fontsize=16, fontweight='bold')
        ax.set_xlabel('X坐标 (堆场宽度方向)', fontsize=12)
        ax.set_ylabel('Y坐标 (堆场长度方向)', fontsize=12)
        ax.grid(True, alpha=0.3)
        ax.set_aspect('equal')

        power_patch = mpatches.Patch(color='lightgreen', label='有电箱区')
        no_power_patch = mpatches.Patch(color='lightcoral', label='无电箱区')
        h_track_patch = mpatches.Patch(color='red', label='横向轨道')
        v_track_patch = mpatches.Patch(color='blue', label='竖向轨道')
        ax.legend(handles=[power_patch, no_power_patch, h_track_patch, v_track_patch],
                  loc='upper left', bbox_to_anchor=(0.02, 0.98))

        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"实际堆场布局图已保存到: {save_path}")
        plt.show()

    def _draw_blocks(self, ax, blocks):
        """绘制箱区"""
        for block_id, block_info in blocks.items():
            x, y = block_info.position_x, block_info.position_y
            color = 'lightgreen' if block_info.has_power else 'lightcoral'

            rect = plt.Rectangle((x - 30, y - 30), 60, 60,
                                 facecolor=color, edgecolor='black', linewidth=1.2)
            ax.add_patch(rect)
            ax.text(x, y, block_id, ha='center', va='center', fontsize=8, fontweight='bold')

            usage_rate = block_info.current_usage / block_info.capacity
            ax.text(x, y - 15, f"{usage_rate:.1%}", ha='center', va='center',
                    fontsize=6, color='darkblue')

    def _draw_tracks(self, ax, tracks):
        """绘制轨道"""
        for track_id, track_info in tracks.items():
            start_x, start_y = track_info.start_point
            end_x, end_y = track_info.end_point

            color = 'red' if track_info.track_type == 'horizontal' else 'blue'
            ax.plot([start_x, end_x], [start_y, end_y], color=color, linewidth=2, alpha=0.8)

    def _add_area_labels(self, ax):
        """添加区域标签"""
        area_centers = {'A': (900, 500), 'B': (650, 800), 'C': (380, 1000), 'D': (180, 500)}
        for area, (x, y) in area_centers.items():
            ax.text(x, y, f"{area}区", ha='center', va='center',
                    fontsize=14, fontweight='bold', color='darkblue',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor="yellow", alpha=0.7))


# ================================
# 5. RTG调度算法（增加三个规则）
# ================================
class AdvancedRTGScheduler:
    """符合实际作业逻辑的RTG调度算法"""

    def __init__(self, config: RealYardConfig):
        self.config = config
        self.tasks = []
        self.rtgs = []
        self.trucks = []
        self.blocks = {}
        self.tracks = {}
        self.track_id_to_name = {}
        self.track_name_to_id = {}
        self.BOX_LAYOUT = {}

        # RTG任务队列管理
        self.rtg_queues: Dict[str, RTGTaskQueue] = {}

        # 三个规则相关变量
        self.urgent_tasks = set()  # 紧急任务集合（规则1：时间承诺制）

    def initialize_data(self, tasks, rtgs, trucks, blocks, tracks, track_id_to_name, track_name_to_id):
        """初始化调度数据"""
        self.tasks = tasks
        self.rtgs = rtgs
        self.trucks = trucks
        self.blocks = blocks
        self.tracks = tracks
        self.track_id_to_name = track_id_to_name
        self.track_name_to_id = track_name_to_id
        self._initialize_box_layout()
        self._initialize_rtg_queues()

    def _initialize_rtg_queues(self):
        """初始化RTG任务队列"""
        for rtg in self.rtgs:
            # 获取RTG初始位置
            initial_block = self._get_rtg_initial_block(rtg)

            self.rtg_queues[rtg.rtg_id] = RTGTaskQueue(
                rtg_id=rtg.rtg_id,
                current_block=initial_block,
                current_bay=rtg.current_bay,
                finish_time=0.0
            )

    def _initialize_box_layout(self):
        """初始化箱区布局"""
        area_blocks = {'A': [], 'B': [], 'C': [], 'D': []}
        for block_id, block_info in self.blocks.items():
            area_blocks[block_info.area].append((block_id, block_info))

        for area, blocks_list in area_blocks.items():
            blocks_list.sort(key=lambda x: x[1].number)
            area_base_col = {'A': 0, 'B': 10, 'C': 20, 'D': 30}[area]

            for i, (block_id, block_info) in enumerate(blocks_list):
                if area == 'A':
                    row, col = i // 4, area_base_col + (i % 4)
                elif area in ['B', 'C']:
                    row, col = i // 4, area_base_col + (i % 4)
                else:  # D区
                    row, col = i // 3, area_base_col + (i % 3)

                self.BOX_LAYOUT[block_id] = {'row': row, 'col': col, 'area': area}

    def calculate_block_distance(self, block1, block2):
        """计算箱区距离"""
        if block1 == block2:
            return 0, 0

        if block1 not in self.BOX_LAYOUT or block2 not in self.BOX_LAYOUT:
            return 0, 0

        row1, col1 = self.BOX_LAYOUT[block1]["row"], self.BOX_LAYOUT[block1]["col"]
        row2, col2 = self.BOX_LAYOUT[block2]["row"], self.BOX_LAYOUT[block2]["col"]

        return abs(row1 - row2), abs(col1 - col2)

    def calculate_move_time(self, current_pos, target_pos, block_current, block_target):
        """计算移动时间"""
        block_row_distance, block_col_distance = self.calculate_block_distance(block_current, block_target)

        if block_row_distance == 0:
            horizontal_move_time = abs(current_pos - target_pos) * 0.28
        else:
            if current_pos < 10 and target_pos < 10:
                horizontal_move_time = (current_pos + target_pos) * 0.28 + 3
            elif current_pos > 10 and target_pos > 10:
                horizontal_move_time = abs(current_pos - 19) * 0.28 + abs(target_pos - 19) * 0.28 + 3
            elif current_pos > 10 and target_pos < 10:
                horizontal_move_time = ((19 - current_pos) + target_pos) * 0.28 + 3
            else:
                horizontal_move_time = (abs(19 - target_pos) + current_pos) * 0.28 + 3

        if block_col_distance == 0 or block_col_distance == 1:
            vertical_move_time = block_col_distance * 10
        else:
            vertical_move_time = 13

        return horizontal_move_time + vertical_move_time + 0.5  # 加上吊具展开的时间30秒

    def _get_rtg_initial_block(self, rtg: RTGInfo) -> str:
        """获取RTG初始所在箱区"""
        # 根据轨道找到对应的箱区
        if hasattr(rtg, 'available_tracks') and rtg.available_tracks:
            # 从可用轨道中找到对应的箱区
            for track_name in rtg.available_tracks:
                if track_name in self.track_name_to_id:
                    track_id = self.track_name_to_id[track_name]
                    if track_id in self.tracks:
                        track = self.tracks[track_id]
                        if track.connected_blocks:
                            return track.connected_blocks[0]

        # 如果找不到，就找第一个该RTG可达的箱区
        for block_id, block_info in self.blocks.items():
            if (block_info.area in rtg.available_areas and
                    self._can_rtg_reach_block(rtg, block_info)):
                return block_id
        # 兜底：返回任意箱区
        return list(self.blocks.keys())[0] if self.blocks else "A01"

    def _can_rtg_reach_block(self, rtg: RTGInfo, block_info: BlockInfo) -> bool:
        """判断RTG是否能到达箱区（严格按照轨道限制）"""
        # 检查区域权限
        if block_info.area not in rtg.available_areas:
            return False
        # 检查设备状态（硬性条件）
        if rtg.login_status != 'online' or rtg.fault_status != 'normal':
            return False
        # 检查轨道可达性
        if not hasattr(rtg, 'available_tracks') or not rtg.available_tracks:
            # 如果没有轨道限制，默认可达
            return True

        # 需要建立T001到实际轨道ID的映射
        rtg_actual_tracks = []
        for t_track in rtg.available_tracks:  # ['T001', 'T002'...]
            # 查找T001对应的实际轨道ID
            if t_track in self.track_name_to_id:
                actual_track_id = self.track_name_to_id[t_track]
                rtg_actual_tracks.append(actual_track_id)
            # 或者直接检查T001是否就是轨道ID
            elif t_track in self.tracks:
                rtg_actual_tracks.append(t_track)

        rtg_tracks_set = set(rtg_actual_tracks)
        block_tracks_set = set(block_info.track_connections)
        return bool(rtg_tracks_set & block_tracks_set)

    def check_task_urgency(self, current_time: float):
        """规则1：时间承诺制 - 检查任务是否违约变为紧急"""
        newly_urgent = []
        for task in self.tasks:
            if (task.deadline and
                    current_time > task.deadline and
                    task.task_id not in self.urgent_tasks):
                self.urgent_tasks.add(task.task_id)
                newly_urgent.append(task.task_id)
                overdue_time = current_time - task.deadline
                print(f"🚨 任务{task.task_id}({task.task_type})已违约{overdue_time:.1f}分钟，设为紧急任务！")

        return len(newly_urgent) > 0

    # ================================
    # 规则2：箱区切换优化策略
    # ================================

    def calculate_block_switch_penalty(self, rtg_assignments: Dict) -> float:
        """
        计算RTG箱区切换惩罚
        rtg_assignments: {rtg_id: [task1, task2, ...]}
        """
        penalty = 0

        for rtg_id, task_list in rtg_assignments.items():
            if len(task_list) <= 1:
                continue

            # 获取RTG处理的所有箱区
            blocks_sequence = [task.block_id for task in task_list]

            # 计算箱区切换次数
            switches = 0
            for i in range(1, len(blocks_sequence)):
                if blocks_sequence[i] != blocks_sequence[i - 1]:
                    switches += 1

            # 箱区切换惩罚：每次切换惩罚300分钟
            penalty += switches * 300

            # 额外惩罚：如果RTG在多个箱区间频繁切换
            unique_blocks = len(set(blocks_sequence))
            if unique_blocks > 2:  # 超过2个箱区
                penalty += (unique_blocks - 2) * 150

        return penalty

    def assign_rtg_by_proximity(self, task: TaskInfo) -> Optional[RTGInfo]:
        """
        基于就近原则分配RTG（优化箱区切换）
        """
        if task.block_id not in self.blocks:
            return None

        block_info = self.blocks[task.block_id]
        available_rtgs = []

        # 筛选可用RTG
        for rtg in self.rtgs:
            if self._can_rtg_reach_block(rtg, block_info):
                available_rtgs.append(rtg)

        if not available_rtgs:
            return None

        # 就近原则选择RTG
        best_rtg = None
        min_cost = float('inf')

        for rtg in available_rtgs:
            cost = self._calculate_proximity_cost(rtg, task)
            if cost < min_cost:
                min_cost = cost
                best_rtg = rtg

        return best_rtg

    def _calculate_proximity_cost(self, rtg: RTGInfo, task: TaskInfo) -> float:
        """
        计算RTG到任务的就近成本（考虑箱区切换）
        """
        # 基础成本：RTG队列等待时间 + 移动时间 + 作业时间
        base_cost = self._calculate_rtg_cost(rtg, task)

        # 箱区切换成本
        rtg_queue = self.rtg_queues[rtg.rtg_id]
        current_block = rtg_queue.current_block if rtg_queue.current_block else task.block_id

        # 如果RTG需要切换箱区，增加成本
        if current_block != task.block_id:
            switch_penalty = 200  # 箱区切换基础惩罚

            # 计算箱区间物理距离
            distance_penalty = self._calculate_block_distance_penalty(current_block, task.block_id)
            base_cost += switch_penalty + distance_penalty
        else:
            # 同箱区作业奖励
            base_cost -= 50

        return base_cost

    def _calculate_block_distance_penalty(self, block1: str, block2: str) -> float:
        """计算箱区间距离惩罚"""
        if block1 == block2:
            return 0

        # 使用BOX_LAYOUT计算距离（如果有的话）
        if block1 in self.BOX_LAYOUT and block2 in self.BOX_LAYOUT:
            row1, col1 = self.BOX_LAYOUT[block1]["row"], self.BOX_LAYOUT[block1]["col"]
            row2, col2 = self.BOX_LAYOUT[block2]["row"], self.BOX_LAYOUT[block2]["col"]

            row_distance = abs(row1 - row2)
            col_distance = abs(col1 - col2)

            # 距离越远，惩罚越大
            return (row_distance + col_distance) * 50
        else:
            # 默认跨箱区惩罚
            return 100

    # ================================
    # 规则3：RTG安全规则和冲突检测
    # ================================

    def detect_rtg_conflicts_in_block(self, block_id: str, task_timings: Dict) -> List[Dict]:
        """
        检测同一箱区内的RTG冲突
        """
        conflicts = []

        # 获取该箱区的所有RTG任务
        block_tasks = []
        for task_id, data in task_timings.items():
            if data['block_id'] == block_id:
                block_tasks.append((task_id, data))

        # 按时间排序
        block_tasks.sort(key=lambda x: x[1]['actual_start_time'])

        # 检测两两之间的冲突
        for i in range(len(block_tasks)):
            for j in range(i + 1, len(block_tasks)):
                task1_id, task1_data = block_tasks[i]
                task2_id, task2_data = block_tasks[j]

                # 检查时间重叠
                if (task1_data['actual_end_time'] > task2_data['actual_start_time'] and
                        task2_data['actual_end_time'] > task1_data['actual_start_time']):

                    # 检查位置冲突
                    conflict_info = self._analyze_rtg_position_conflict(task1_data, task2_data)
                    if conflict_info['has_conflict']:
                        conflicts.append({
                            'task1': task1_id,
                            'task2': task2_id,
                            'conflict_type': conflict_info['type'],
                            'severity': conflict_info['severity'],
                            'resolution': self._resolve_rtg_conflict(task1_data, task2_data)
                        })

        return conflicts

    def _analyze_rtg_position_conflict(self, task1_data: Dict, task2_data: Dict) -> Dict:
        """
        分析RTG位置冲突（改进版安全规则）
        """
        rtg1_pos = task1_data['bay']
        rtg2_pos = task2_data['bay']
        rtg1_current = task1_data.get('current_bay', rtg1_pos)
        rtg2_current = task2_data.get('current_bay', rtg2_pos)

        # 检查是否携带箱子
        rtg1_has_container = task1_data.get('has_container', False)
        rtg2_has_container = task2_data.get('has_container', False)
        rtg1_container_size = task1_data.get('container_size', 40)  # 默认40寸
        rtg2_container_size = task2_data.get('container_size', 40)

        # 确定安全距离
        safety_distance = self._calculate_safety_distance(
            rtg1_has_container, rtg1_container_size,
            rtg2_has_container, rtg2_container_size
        )

        conflict_info = {
            'has_conflict': False,
            'type': None,
            'severity': 0,
            'safety_distance_required': safety_distance
        }

        # 1. 检查路径交叉冲突
        if self._check_path_crossing(rtg1_current, rtg1_pos, rtg2_current, rtg2_pos):
            conflict_info.update({
                'has_conflict': True,
                'type': 'path_crossing',
                'severity': 9
            })
            return conflict_info

        # 2. 检查位置冲突（考虑安全距离）
        position_distance = abs(rtg1_pos - rtg2_pos)
        if position_distance < safety_distance:
            conflict_info.update({
                'has_conflict': True,
                'type': 'position_conflict',
                'severity': 8
            })
            return conflict_info

        # 3. 检查通过冲突
        if (rtg1_current < rtg2_pos < rtg1_pos or rtg1_pos < rtg2_pos < rtg1_current or
                rtg2_current < rtg1_pos < rtg2_pos or rtg2_pos < rtg1_pos < rtg2_current):
            conflict_info.update({
                'has_conflict': True,
                'type': 'pass_through_conflict',
                'severity': 7
            })

        return conflict_info

    def _calculate_safety_distance(self, rtg1_has_container: bool, rtg1_size: int,
                                   rtg2_has_container: bool, rtg2_size: int) -> int:
        """
        计算RTG间安全距离
        """
        # 基础安全距离
        base_distance = 1

        # 如果任一RTG携带20寸小箱子，需要额外安全距离
        if ((rtg1_has_container and rtg1_size == 20) or
                (rtg2_has_container and rtg2_size == 20)):
            return base_distance + 1  # 保持一个贝位的安全距离

        # 其他情况可以相邻作业
        return base_distance

    def _check_path_crossing(self, start1: int, end1: int, start2: int, end2: int) -> bool:
        """
        检查两个RTG的路径是否交叉
        """
        # 检查路径是否有交叉
        return ((start1 < start2 < end1) or (start1 < end2 < end1) or
                (start2 < start1 < end2) or (start2 < end1 < end2) or
                (start1 < end2 and end1 > start2))

    def _resolve_rtg_conflict(self, task1_data: Dict, task2_data: Dict) -> Dict:
        """
        改进的RTG冲突解决策略
        """
        # 获取任务优先级
        task1_priority = self.config.TASK_PRIORITY.get(task1_data.get('task_type', '进箱'), 5)
        task2_priority = self.config.TASK_PRIORITY.get(task2_data.get('task_type', '进箱'), 5)

        # 规则1：低优先级任务RTG避让
        if task1_priority < task2_priority:
            return {
                'action': 'task2_wait',
                'wait_time': task1_data['actual_end_time'] - task2_data['actual_start_time'],
                'reason': 'task1_higher_priority'
            }
        elif task2_priority < task1_priority:
            return {
                'action': 'task1_wait',
                'wait_time': task2_data['actual_end_time'] - task1_data['actual_start_time'],
                'reason': 'task2_higher_priority'
            }

        # 规则2：优先级相同时，拖车到达时间晚的避让
        task1_truck_arrival = task1_data.get('truck_arrival_time', task1_data.get('prep_time', 0))
        task2_truck_arrival = task2_data.get('truck_arrival_time', task2_data.get('prep_time', 0))

        if task1_truck_arrival <= task2_truck_arrival:
            return {
                'action': 'task2_wait',
                'wait_time': task1_data['actual_end_time'] - task2_data['actual_start_time'],
                'reason': 'task1_earlier_truck_arrival'
            }
        else:
            return {
                'action': 'task1_wait',
                'wait_time': task2_data['actual_end_time'] - task1_data['actual_start_time'],
                'reason': 'task2_earlier_truck_arrival'
            }

    def _calculate_rtg_cost(self, rtg: RTGInfo, task: TaskInfo) -> float:
        """
        计算RTG执行任务的总成本（集成三个规则）
        """
        rtg_queue = self.rtg_queues[rtg.rtg_id]
        # 1. RTG完成当前所有任务的时间
        queue_finish_time = rtg_queue.get_next_available_time()
        # 2. RTG当前位置（完成最后一个任务后的位置）
        current_block, current_bay = rtg_queue.get_current_location()
        # 3. 从当前位置移动到任务位置的时间
        move_time = self.calculate_move_time(
            current_bay, task.bay, current_block, task.block_id
        )
        # 4. 任务执行时间
        task_duration = task.estimated_duration
        # 5. 总成本 = 等待时间 + 移动时间 + 执行时间
        total_cost = max(queue_finish_time, task.truck_arrival_time) + move_time + task_duration

        # 规则1：时间承诺制 - 违约惩罚
        if task.task_id in self.urgent_tasks:
            total_cost -= 500  # 紧急任务最高优先级
        elif task.deadline and total_cost > task.deadline:
            delay = total_cost - task.deadline
            total_cost += delay * 100  # 预期违约的惩罚

        return total_cost

    def assign_rtg_to_task(self, task: TaskInfo) -> Optional[RTGInfo]:
        """为任务分配最优RTG（使用就近原则）"""
        # 使用规则2的就近原则分配
        return self.assign_rtg_by_proximity(task)

    def assign_task_to_rtg(self, task: TaskInfo, rtg: RTGInfo) -> Dict:
        """将任务分配给指定RTG并更新队列"""
        rtg_queue = self.rtg_queues[rtg.rtg_id]
        # 计算任务开始时间（等待队列完成和拖车到达）
        start_time = max(rtg_queue.get_next_available_time(), task.truck_arrival_time)
        # 计算移动时间
        current_block, current_bay = rtg_queue.get_current_location()
        move_time = self.calculate_move_time(
            current_bay, task.bay, current_block, task.block_id
        )
        # 实际开始作业时间 = 等待时间 + 移动时间
        actual_start_time = start_time + move_time
        end_time = actual_start_time + task.estimated_duration
        # 创建任务分配记录
        assignment = {
            'task_id': task.task_id,
            'rtg_id': rtg.rtg_id,
            'block_id': task.block_id,
            'bay': task.bay,
            'row': task.row,
            'tier': task.tier,
            'task_type': task.task_type,
            'estimated_time': task.estimated_duration,
            'is_cold': task.is_cold_container,
            'manual_mark': task.manual_mark,
            'truck_id': task.truck_id,
            'deadline': task.deadline,
            'container_size': task.container_size,
            'has_container': getattr(rtg, 'has_container', False),
            'truck_arrival_time': task.truck_arrival_time,
            'queue_wait_time': start_time,
            'move_time': move_time,
            'start_time': actual_start_time,
            'end_time': end_time,
            'is_urgent': task.task_id in self.urgent_tasks,
            'priority_level': 0 if task.manual_mark else self.config.TASK_PRIORITY.get(task.task_type, 99)
            # manual_mark任务最高优先级
        }
        # 更新RTG队列
        rtg_queue.add_task(assignment)
        return assignment

    def preprocess_tasks(self):
        """任务预处理"""
        manual_tasks = [task for task in self.tasks if task.manual_mark]
        normal_tasks = [task for task in self.tasks if not task.manual_mark]

        manual_tasks.sort(key=lambda t: (0 if t.is_cold_container else 1))

        def task_sort_key(task):
            # 紧急任务优先
            urgency_priority = 0 if task.task_id in self.urgent_tasks else 1
            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            cold_priority = 0 if task.is_cold_container else 1
            return (urgency_priority, task_priority, cold_priority, task.truck_arrival_time)

        normal_tasks.sort(key=task_sort_key)
        return manual_tasks, normal_tasks

    def optimize_schedule(self):
        """优化调度方案（集成三个规则）"""
        print("\n正在优化RTG调度方案（集成三个规则）...")
        manual_tasks, normal_tasks = self.preprocess_tasks()

        schedule = {
            'manual_task_assignments': [],
            'normal_task_assignments': [],
            'total_makespan': 0,
            'rtg_utilization': {},
            'priority_summary': {},
            'rtg_queues': {},  # 记录每个RTG的任务队列
            'urgent_tasks_count': 0,  # 紧急任务数量
            'time_commitment_summary': {},  # 时间承诺摘要
            'block_switch_penalty': 0,  # 箱区切换惩罚
            'safety_conflicts': [],  # 安全冲突
            'rtg_assignments': {rtg.rtg_id: [] for rtg in self.rtgs}  # RTG任务分配
        }

        total_tasks = len(self.tasks)
        assigned_count = 0
        current_time = 0

        # 处理人工标记任务
        print(f"处理人工标记任务 ({len(manual_tasks)}个):")
        for i, task in enumerate(manual_tasks):
            # 检查时间承诺违约情况
            self.check_task_urgency(current_time)

            rtg = self.assign_rtg_to_task(task)
            if rtg:
                assignment = self.assign_task_to_rtg(task, rtg)
                schedule['manual_task_assignments'].append(assignment)
                schedule['rtg_assignments'][rtg.rtg_id].append(task)
                assigned_count += 1
                current_time = max(current_time, assignment['end_time'])

                cold_mark = "(冷箱)" if task.is_cold_container else ""
                urgent_mark = "(🚨紧急)" if task.task_id in self.urgent_tasks else ""
                print(f"  {i + 1}. {task.task_id}: {task.task_type}{cold_mark}{urgent_mark}")
            else:
                print(f"  ❌ 任务{task.task_id}分配失败")
        # 处理普通任务
        print(f"处理普通任务 ({len(normal_tasks)}个):")
        current_priority = None
        task_count = 0
        for task in normal_tasks:
            # 动态检查时间承诺违约
            self.check_task_urgency(current_time)

            task_priority = self.config.TASK_PRIORITY.get(task.task_type, 99)
            if current_priority != task_priority:
                if current_priority is not None:
                    print(f"  优先级{current_priority}完成，共{task_count}个任务")
                current_priority = task_priority
                task_count = 0
                priority_names = {1: "装卸类", 2: "收发类", 3: "翻倒类"}
                print(f"  开始处理优先级{task_priority}({priority_names.get(task_priority, '其他')}):")
            rtg = self.assign_rtg_to_task(task)
            if rtg:
                assignment = self.assign_task_to_rtg(task, rtg)
                schedule['normal_task_assignments'].append(assignment)
                schedule['rtg_assignments'][rtg.rtg_id].append(task)
                assigned_count += 1
                task_count += 1
                current_time = max(current_time, assignment['end_time'])

                # 显示时间承诺状态
                status_msg = ""
                if task.deadline:
                    if assignment['end_time'] <= task.deadline:
                        status_msg = "(✅按时)"
                    else:
                        delay = assignment['end_time'] - task.deadline
                        status_msg = f"(⚠️延迟{delay:.1f}min)"

                urgent_mark = "(🚨紧急)" if task.task_id in self.urgent_tasks else ""
                print(f"    {task.task_id}: {task.task_type}{urgent_mark}{status_msg}")
            else:
                print(f"    ❌ 任务{task.task_id}分配失败")
        if current_priority is not None:
            print(f"  优先级{current_priority}完成，共{task_count}个任务")

        # 计算箱区切换惩罚（规则2）
        schedule['block_switch_penalty'] = self.calculate_block_switch_penalty(
            schedule['rtg_assignments']
        )

        # 检测安全冲突（规则3）
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']
        task_timings = {}

        for assignment in all_assignments:
            task_timings[assignment['task_id']] = {
                'block_id': assignment['block_id'],
                'bay': assignment['bay'],
                'actual_start_time': assignment['start_time'],
                'actual_end_time': assignment['end_time'],
                'task_type': assignment.get('task_type', '进箱'),
                'prep_time': assignment.get('truck_arrival_time', 0),
                'has_container': assignment.get('has_container', False),
                'container_size': assignment.get('container_size', 40),
                'truck_arrival_time': assignment.get('truck_arrival_time', 0)
            }

        # 按箱区检测冲突
        all_blocks = set(data['block_id'] for data in task_timings.values())
        for block_id in all_blocks:
            block_conflicts = self.detect_rtg_conflicts_in_block(block_id, task_timings)
            schedule['safety_conflicts'].extend(block_conflicts)

        # 计算最终结果
        if all_assignments:
            schedule['total_makespan'] = max(assignment['end_time'] for assignment in all_assignments)
        else:
            schedule['total_makespan'] = 0
        # 计算RTG利用率
        rtg_workload = {}
        for rtg_id, queue in self.rtg_queues.items():
            rtg_workload[rtg_id] = queue.finish_time
        schedule['rtg_utilization'] = rtg_workload

        # 计算时间承诺统计
        self._calculate_time_commitment_summary(schedule)

        # 保存RTG队列详情
        schedule['rtg_queues'] = {
            rtg_id: {
                'task_count': len(queue.tasks),
                'total_time': queue.finish_time,
                'tasks': queue.tasks
            }
            for rtg_id, queue in self.rtg_queues.items()
        }

        print(f"\n调度优化完成:")
        print(f"  - 总任务数: {total_tasks}")
        print(f"  - 成功分配: {assigned_count}个 ({assigned_count / total_tasks * 100:.1f}%)")
        print(f"  - 预计总时间: {schedule['total_makespan']:.1f}分钟")
        print(f"  - 紧急任务数: {len(self.urgent_tasks)}个")
        print(f"  - 箱区切换惩罚: {schedule['block_switch_penalty']:.1f}")
        print(f"  - 安全冲突数量: {len(schedule['safety_conflicts'])}")

        # 显示RTG箱区分配统计
        self._print_rtg_block_assignments(schedule['rtg_assignments'])

        # 显示安全冲突详情
        if schedule['safety_conflicts']:
            self._print_safety_conflicts(schedule['safety_conflicts'])

        return schedule

    def _print_rtg_block_assignments(self, rtg_assignments: Dict):
        """打印RTG箱区分配统计"""
        print(f"\n=== RTG箱区分配统计 ===")

        for rtg_id, task_list in rtg_assignments.items():
            if not task_list:
                continue

            blocks = [task.block_id for task in task_list]
            unique_blocks = list(set(blocks))
            switches = sum(1 for i in range(1, len(blocks)) if blocks[i] != blocks[i - 1])

            print(f"RTG-{rtg_id}:")
            print(f"  - 任务数: {len(task_list)}")
            print(f"  - 涉及箱区: {unique_blocks}")
            print(f"  - 箱区切换次数: {switches}")

    def _print_safety_conflicts(self, conflicts: List[Dict]):
        """打印安全冲突详情"""
        print(f"\n=== 安全冲突详情 ===")

        for i, conflict in enumerate(conflicts):
            print(f"冲突{i + 1}:")
            print(f"  - 任务: {conflict['task1']} vs {conflict['task2']}")
            print(f"  - 类型: {conflict['conflict_type']}")
            print(f"  - 严重程度: {conflict['severity']}")
            print(f"  - 解决方案: {conflict['resolution']['action']}")
            print(f"  - 原因: {conflict['resolution']['reason']}")

    def _calculate_time_commitment_summary(self, schedule: Dict):
        """计算时间承诺统计"""
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']

        commitment_stats = {
            'total_commitment_tasks': 0,
            'on_time_tasks': 0,
            'delayed_tasks': 0,
            'urgent_tasks': len(self.urgent_tasks),
            'average_delay': 0
        }

        delays = []

        for assignment in all_assignments:
            task_id = assignment['task_id']
            task = next((t for t in self.tasks if t.task_id == task_id), None)

            if task and task.deadline:
                commitment_stats['total_commitment_tasks'] += 1
                actual_completion = assignment['end_time']

                if actual_completion <= task.deadline:
                    commitment_stats['on_time_tasks'] += 1
                else:
                    commitment_stats['delayed_tasks'] += 1
                    delays.append(actual_completion - task.deadline)

        if delays:
            commitment_stats['average_delay'] = np.mean(delays)

        schedule['time_commitment_summary'] = commitment_stats
        schedule['urgent_tasks_count'] = len(self.urgent_tasks)

        print(f"\n=== 时间承诺执行情况 ===")
        print(f"有deadline任务总数: {commitment_stats['total_commitment_tasks']}")
        print(f"按时完成: {commitment_stats['on_time_tasks']}")
        print(f"延迟任务: {commitment_stats['delayed_tasks']}")
        print(f"紧急任务: {commitment_stats['urgent_tasks']}")
        if commitment_stats['average_delay'] > 0:
            print(f"平均延迟: {commitment_stats['average_delay']:.1f}分钟")

    def export_schedule_to_excel(self, schedule, filename=None):
        """导出调度结果到Excel"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"RTG调度结果_三规则_{timestamp}.xlsx"

        print(f"正在导出调度结果到: {filename}")

        wb = Workbook()
        wb.remove(wb.active)

        # 详细调度结果
        ws_detail = wb.create_sheet("详细调度结果")
        all_assignments = schedule['manual_task_assignments'] + schedule['normal_task_assignments']

        headers = ['任务ID', '箱区', '贝位', '排', '层', '任务类型', '分配RTG', 'RTG类型',
                   '拖车到达时间', '开始时间(分钟)', '结束时间(分钟)', '作业时长(分钟)',
                   'deadline', '是否冷箱', '是否紧急', '优先级']
        ws_detail.append(headers)

        for assignment in all_assignments:
            task_detail = next((t for t in self.tasks if t.task_id == assignment['task_id']), None)
            rtg_detail = next((r for r in self.rtgs if r.rtg_id == assignment['rtg_id']), None)

            # 格式化时间显示 - 从assignment中读取
            truck_arrival_display = ""
            deadline_display = ""

            # 从assignment中获取拖车到达时间
            truck_arrival_time = assignment.get('truck_arrival_time', 0)
            if truck_arrival_time and truck_arrival_time > 0:
                hours = int(truck_arrival_time // 60)
                minutes = int(truck_arrival_time % 60)
                truck_arrival_display = f"{hours:02d}:{minutes:02d}"

            # 从assignment中获取deadline
            deadline = assignment.get('deadline', None)
            if deadline and deadline > 0:
                hours = int(deadline // 60)
                minutes = int(deadline % 60)
                deadline_display = f"{hours:02d}:{minutes:02d}"

            row = [
                assignment['task_id'],
                assignment['block_id'],
                task_detail.bay if task_detail else '',
                task_detail.row if task_detail else '',
                task_detail.tier if task_detail else '',
                assignment['task_type'],
                assignment['rtg_id'],
                rtg_detail.rtg_type if rtg_detail else '',
                truck_arrival_display,  # 格式化的拖车到达时间
                round(assignment['start_time'], 2),
                round(assignment['end_time'], 2),
                round(assignment['estimated_time'], 2),
                deadline_display,  # 格式化的deadline
                '是' if assignment['is_cold'] else '否',
                '是' if assignment['is_urgent'] else '否',
                assignment['priority_level']
            ]
            ws_detail.append(row)

        # RTG工作负载统计
        ws_rtg = wb.create_sheet("RTG工作负载")
        ws_rtg.append(['RTG编号', 'RTG类型', '任务数量', '总工作时间', '利用率(%)', '涉及箱区数', '箱区切换次数'])

        rtg_stats = {}
        for assignment in all_assignments:
            rtg_id = assignment['rtg_id']
            if rtg_id not in rtg_stats:
                rtg_info = next((r for r in self.rtgs if r.rtg_id == rtg_id), None)
                rtg_stats[rtg_id] = {
                    'type': rtg_info.rtg_type if rtg_info else 'unknown',
                    'tasks': 0,
                    'time': 0,
                    'blocks': []
                }
            rtg_stats[rtg_id]['tasks'] += 1
            rtg_stats[rtg_id]['time'] += assignment['estimated_time']
            rtg_stats[rtg_id]['blocks'].append(assignment['block_id'])

        total_makespan = schedule['total_makespan']
        for rtg_id, stats in rtg_stats.items():
            utilization = stats['time'] / total_makespan * 100 if total_makespan > 0 else 0
            blocks = stats['blocks']
            unique_blocks = len(set(blocks))
            switches = sum(1 for i in range(1, len(blocks)) if blocks[i] != blocks[i - 1])

            ws_rtg.append([
                rtg_id,
                stats['type'],
                stats['tasks'],
                round(stats['time'], 2),
                round(utilization, 2),
                unique_blocks,
                switches
            ])

        # 三规则摘要
        ws_summary = wb.create_sheet("三规则摘要")
        summary = schedule['time_commitment_summary']
        ws_summary.append(['指标', '数值'])
        ws_summary.append(['规则1: 时间承诺制任务数', summary['total_commitment_tasks']])
        ws_summary.append(['按时完成任务', summary['on_time_tasks']])
        ws_summary.append(['延迟任务', summary['delayed_tasks']])
        ws_summary.append(['紧急任务', summary['urgent_tasks']])
        ws_summary.append(['平均延迟(分钟)', round(summary['average_delay'], 2)])
        ws_summary.append(['规则2: 箱区切换总惩罚', round(schedule['block_switch_penalty'], 2)])
        ws_summary.append(['规则3: 安全冲突数量', len(schedule['safety_conflicts'])])
        ws_summary.append(['总完成时间(分钟)', round(schedule['total_makespan'], 2)])

        # 安全冲突详情
        if schedule['safety_conflicts']:
            ws_conflicts = wb.create_sheet("安全冲突详情")
            ws_conflicts.append(['冲突编号', '任务1', '任务2', '冲突类型', '严重程度', '解决方案', '原因'])

            for i, conflict in enumerate(schedule['safety_conflicts']):
                ws_conflicts.append([
                    i + 1,
                    conflict['task1'],
                    conflict['task2'],
                    conflict['conflict_type'],
                    conflict['severity'],
                    conflict['resolution']['action'],
                    conflict['resolution']['reason']
                ])

        wb.save(filename)
        print(f"调度结果已导出到: {filename}")
        return filename


# ================================
# 6. 主函数
# ================================
def main():
    """主函数"""
    print("=== RTG调度系统（集成规则） ===")
    config = RealYardConfig()
    data_loader = YardDataLoader(config)

    # 加载数据
    blocks, tracks, track_id_to_name, track_name_to_id = data_loader.load_yard_layout()
    tasks = data_loader.load_tasks_from_excel()
    rtgs = data_loader.load_rtgs_from_excel()
    trucks = data_loader.load_trucks(tasks)

    print(f"\n数据加载完成:")
    print(f"  - 任务数量: {len(tasks)}")
    print(f"  - RTG数量: {len(rtgs)}")
    print(f"  - 箱区数量: {len(blocks)}")
    print(f"  - 轨道数量: {len(tracks)}")

    # 统计时间承诺任务
    deadline_tasks = [t for t in tasks if t.deadline]
    in_out_tasks = [t for t in tasks if t.task_type in ["进箱", "提箱"]]
    print(f"  - 有deadline任务: {len(deadline_tasks)}个")
    print(f"  - 进箱/提箱任务: {len(in_out_tasks)}个")

    # 可视化
    visualizer = RealYardVisualization(config)
    visualizer.plot_yard_layout(blocks, tracks)

    print("=" * 50)
    print("运行集成规则RTG调度算法...")

    # 运行队列调度算法
    scheduler = AdvancedRTGScheduler(config)
    scheduler.initialize_data(tasks, rtgs, trucks, blocks, tracks,
                              track_id_to_name, track_name_to_id)
    schedule = scheduler.optimize_schedule()

    # 显示结果
    print(f"\n=== 调度结果摘要 ===")
    print(f"人工标记任务: {len(schedule['manual_task_assignments'])}个")
    print(f"普通任务分配: {len(schedule['normal_task_assignments'])}个")
    print(f"预计完成时间: {schedule['total_makespan']:.1f}分钟")
    print(f"紧急任务数: {schedule['urgent_tasks_count']}个")

    print(f"\n=== RTG工作负载分布 ===")
    rtg_items = sorted(schedule['rtg_utilization'].items(), key=lambda x: x[1], reverse=True)
    for rtg_id, workload in rtg_items[:8]:
        rtg_info = next((r for r in rtgs if r.rtg_id == rtg_id), None)
        if rtg_info and workload > 0:
            utilization = workload / schedule['total_makespan'] * 100
            print(f"  {rtg_id}({rtg_info.rtg_type}): {workload:.1f}分钟 (利用率{utilization:.1f}%)")

    print(f"\n系统已完成集成三规则的RTG调度优化！")

    # 导出Excel
    try:
        excel_file = scheduler.export_schedule_to_excel(schedule)
        print(f"调度结果已导出到Excel: {excel_file}")
    except Exception as e:
        print(f"Excel导出失败: {e}")
    return tasks, rtgs, trucks, blocks, tracks, schedule


if __name__ == "__main__":
    tasks, rtgs, trucks, blocks, tracks, schedule = main()